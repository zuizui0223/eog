from __future__ import annotations

import hashlib
import json
import math
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
HERE = Path(__file__).resolve().parent
BUILD = ROOT / "build" / "bodie_pika_replication_2"
BUILD.mkdir(parents=True, exist_ok=True)
CONTRACT = json.loads((HERE / "source_contract.json").read_text())

DOI = CONTRACT["source"]["doi"]
RESPONSE_NAME = CONTRACT["source"]["response_bearing_filename"]
ALLOWED = set(CONTRACT["source"]["allowed_non_response_filenames"])
EXPECTED_N = int(CONTRACT["geometry_gate"]["expected_complete_patch_registry_count"])
TARGETS = [float(x) for x in CONTRACT["geometry_gate"]["structural_lcc_targets"]]
MIN_SCALES = int(CONTRACT["geometry_gate"]["minimum_distinct_positive_thresholds"])
R_KM = float(CONTRACT["geometry_gate"]["haversine_radius_km"])
API = "https://datadryad.org/api/v2"
UA = "EOG-Bodie-Gate0/1.0 (+response-free metadata and non-census geometry audit)"


def canonical_sha256(obj: Any) -> str:
    raw = json.dumps(obj, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode()
    return hashlib.sha256(raw).hexdigest()


def get_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "X-API-Version": "2.1.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()


def get_json(url: str) -> dict[str, Any]:
    return json.loads(get_bytes(url))


def api_url(href: str) -> str:
    if href.startswith("http://") or href.startswith("https://"):
        return href
    return "https://datadryad.org" + href


def extract_id(file_obj: dict[str, Any]) -> int:
    href = file_obj.get("_links", {}).get("self", {}).get("href", "")
    m = re.search(r"/files/(\d+)$", href)
    if not m:
        raise RuntimeError(f"cannot extract Dryad file id from {href!r}")
    return int(m.group(1))


def latest_version() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    encoded = urllib.parse.quote(f"doi:{DOI}", safe="")
    ds_url = f"{API}/datasets/{encoded}"
    ds = get_json(ds_url)
    versions = get_json(f"{ds_url}/versions")
    vv = versions.get("_embedded", {}).get("stash:versions", [])
    if not vv:
        raise RuntimeError("Dryad versions list is empty")
    latest = vv[-1]
    files_href = latest.get("_links", {}).get("stash:files", {}).get("href")
    version_id = latest.get("id")
    if files_href:
        files_json = get_json(api_url(files_href))
    elif version_id is not None:
        files_json = get_json(f"{API}/versions/{version_id}/files")
    else:
        raise RuntimeError("latest Dryad version lacks files link/id")
    files = files_json.get("_embedded", {}).get("stash:files", [])
    if not files:
        raise RuntimeError("latest Dryad version file inventory is empty")
    return {"dataset": ds, "version": latest}, files


def file_meta(f: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": extract_id(f),
        "path": f.get("path"),
        "size": f.get("size"),
        "mimeType": f.get("mimeType"),
        "digest": f.get("digest"),
        "digestType": f.get("digestType"),
    }


def download_allowed(f: dict[str, Any]) -> tuple[Path, dict[str, Any]]:
    meta = file_meta(f)
    name = str(meta["path"])
    if name not in ALLOWED:
        raise RuntimeError(f"attempted non-response download outside allowlist: {name}")
    fid = int(meta["id"])
    url = f"https://datadryad.org/stash/downloads/file_stream/{fid}"
    data = get_bytes(url)
    if meta["size"] is not None and len(data) != int(meta["size"]):
        raise RuntimeError(f"size mismatch for {name}: {len(data)} != {meta['size']}")
    dtype = str(meta.get("digestType") or "").lower().replace("_", "-")
    expected = str(meta.get("digest") or "").lower()
    if expected:
        if dtype in {"sha-256", "sha256"}:
            actual = hashlib.sha256(data).hexdigest()
        elif dtype in {"md5"}:
            actual = hashlib.md5(data).hexdigest()
        else:
            raise RuntimeError(f"unsupported Dryad digest type for allowed file {name}: {dtype}")
        if actual != expected:
            raise RuntimeError(f"digest mismatch for {name}")
    else:
        actual = hashlib.sha256(data).hexdigest()
        dtype = "locally-computed-sha-256"
    out = BUILD / name
    out.write_bytes(data)
    return out, {**meta, "downloaded_bytes": len(data), "verified_digest": actual, "verified_digest_type": dtype}


def norm(v: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(v or "").strip().lower())


def numeric(v: Any) -> float | None:
    if isinstance(v, bool) or v is None:
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    return x if math.isfinite(x) else None


ID_NAMES = {"patch", "patchid", "patchnumber", "patchno", "site", "siteid", "habitatpatch", "oredump", "dump", "dumpid"}
LAT_NAMES = {"lat", "latitude", "decimallatitude"}
LON_NAMES = {"lon", "long", "longitude", "decimallongitude"}
E_NAMES = {"easting", "utmeasting", "utmx", "xutm"}
N_NAMES = {"northing", "utmnorthing", "utmy", "yutm"}


def scan_geometry_workbook(path: Path, readme_text: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    diagnostics: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    readme_low = readme_text.lower()
    for ws in wb.worksheets:
        preview = list(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True))
        diagnostics.append({
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview_headers_only": [[None if v is None else str(v)[:120] for v in row] for row in preview[:5]],
        })
        for idx, row in enumerate(preview, start=1):
            names = [norm(v) for v in row]
            id_cols = [i for i, x in enumerate(names) if x in ID_NAMES]
            lat_cols = [i for i, x in enumerate(names) if x in LAT_NAMES]
            lon_cols = [i for i, x in enumerate(names) if x in LON_NAMES]
            east_cols = [i for i, x in enumerate(names) if x in E_NAMES]
            north_cols = [i for i, x in enumerate(names) if x in N_NAMES]
            modes: list[tuple[str, int, int]] = []
            if lat_cols and lon_cols:
                modes.append(("latitude_longitude", lat_cols[0], lon_cols[0]))
            if east_cols and north_cols and "utm" in readme_low:
                modes.append(("documented_utm", east_cols[0], north_cols[0]))
            for mode, ci, cj in modes:
                if not id_cols:
                    continue
                ii = id_cols[0]
                records: list[tuple[str, float, float]] = []
                for rr in ws.iter_rows(min_row=idx + 1, values_only=True):
                    if ii >= len(rr) or ci >= len(rr) or cj >= len(rr):
                        continue
                    pid = str(rr[ii]).strip() if rr[ii] is not None else ""
                    x = numeric(rr[ci]); y = numeric(rr[cj])
                    if not pid or x is None or y is None:
                        continue
                    records.append((pid, x, y))
                ids = [r[0] for r in records]
                unique = {x for x in ids}
                duplicate_ids = len(ids) - len(unique)
                semantics_ok = mode == "latitude_longitude" or (mode == "documented_utm" and "utm" in readme_low)
                if mode == "latitude_longitude":
                    range_ok = all(-90 <= r[1] <= 90 and -180 <= r[2] <= 180 for r in records)
                else:
                    range_ok = all(abs(r[1]) < 2e7 and abs(r[2]) < 2e7 for r in records)
                cand = {
                    "sheet": ws.title,
                    "header_row": idx,
                    "id_header": row[ii],
                    "coord_headers": [row[ci], row[cj]],
                    "mode": mode,
                    "record_count": len(records),
                    "unique_patch_count": len(unique),
                    "duplicate_patch_rows": duplicate_ids,
                    "coordinate_semantics_documented": semantics_ok,
                    "coordinate_range_ok": range_ok,
                    "records": records,
                }
                candidates.append(cand)
    wb.close()
    valid = [
        c for c in candidates
        if c["unique_patch_count"] == EXPECTED_N
        and c["record_count"] == EXPECTED_N
        and c["duplicate_patch_rows"] == 0
        and c["coordinate_semantics_documented"]
        and c["coordinate_range_ok"]
    ]
    return {
        "diagnostics": diagnostics,
        "candidate_summaries": [{k: v for k, v in c.items() if k != "records"} for c in candidates],
        "valid": valid,
    }


def haversine_km(a: tuple[float, float], b: tuple[float, float]) -> float:
    lat1, lon1 = map(math.radians, a); lat2, lon2 = map(math.radians, b)
    dlat = lat2 - lat1; dlon = lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * R_KM * math.asin(min(1.0, math.sqrt(h)))


def euclidean_km(a: tuple[float, float], b: tuple[float, float]) -> float:
    return math.hypot(a[0] - b[0], a[1] - b[1]) / 1000.0


def lcc_fraction(n: int, edges: list[tuple[int, int, float]], threshold: float) -> float:
    parent = list(range(n)); size = [1] * n
    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        if size[ra] < size[rb]:
            ra, rb = rb, ra
        parent[rb] = ra; size[ra] += size[rb]
    for i, j, d in edges:
        if d <= threshold:
            union(i, j)
    return max(size[find(i)] for i in range(n)) / n


def ladder(records: list[tuple[str, float, float]], mode: str) -> tuple[list[dict[str, Any]], int]:
    coords = [(r[1], r[2]) for r in records]
    dist = haversine_km if mode == "latitude_longitude" else euclidean_km
    edges: list[tuple[int, int, float]] = []
    vals: list[float] = []
    for i in range(len(coords)):
        for j in range(i + 1, len(coords)):
            d = dist(coords[i], coords[j])
            if d > 0:
                edges.append((i, j, d)); vals.append(d)
    uniq = sorted(set(vals))
    out = []
    for target in TARGETS:
        chosen = None
        for d in uniq:
            frac = lcc_fraction(len(coords), edges, d)
            if frac >= target:
                chosen = (d, frac); break
        if chosen is None:
            raise RuntimeError(f"no threshold reaches LCC target {target}")
        out.append({"target_lcc_fraction": target, "threshold_km": chosen[0], "achieved_lcc_fraction": chosen[1]})
    distinct = len({round(x["threshold_km"], 12) for x in out if x["threshold_km"] > 0})
    return out, distinct


def base_result() -> dict[str, Any]:
    return {
        "schema": "eog.bodie_pika_replication_2.gate0.v1",
        "attempt_id": CONTRACT["attempt_id"],
        "status": "engineering_failure_pre_response",
        "reason": None,
        "dryad": {},
        "non_response_downloads": [],
        "geometry": {},
        "response_firewall": {
            "census_payload_requests": 0,
            "census_payload_bytes_opened": 0,
            "census_header_bytes_opened": 0,
            "census_sheet_names_opened": false,
            "census_rows_opened": false,
            "census_values_opened": false,
            "scientific_model_fits": 0,
            "heldout_scores": 0
        }
    }


def main() -> int:
    result = base_result()
    try:
        meta, files = latest_version()
        metas = [file_meta(f) for f in files]
        by_name = {str(m["path"]): (f, m) for f, m in zip(files, metas)}
        if RESPONSE_NAME not in by_name:
            raise RuntimeError(f"frozen census filename missing from latest Dryad inventory: {RESPONSE_NAME}")
        missing_allowed = sorted(ALLOWED - set(by_name))
        if missing_allowed:
            raise RuntimeError(f"allowed non-response files missing from Dryad inventory: {missing_allowed}")
        census_meta = by_name[RESPONSE_NAME][1]
        result["dryad"] = {
            "dataset_identifier": meta["dataset"].get("identifier"),
            "dataset_title": meta["dataset"].get("title"),
            "dataset_version_number": meta["version"].get("versionNumber"),
            "dataset_last_modification_date": meta["version"].get("lastModificationDate"),
            "version_id": meta["version"].get("id"),
            "file_count": len(metas),
            "census_file_metadata_only": census_meta,
            "inventory_fingerprint": canonical_sha256(metas)
        }
        allowed_paths: dict[str, Path] = {}
        for name in sorted(ALLOWED):
            path, audit = download_allowed(by_name[name][0])
            allowed_paths[name] = path
            result["non_response_downloads"].append(audit)
        readme_text = allowed_paths["README.csv"].read_text(encoding="utf-8-sig", errors="replace")
        scan = scan_geometry_workbook(allowed_paths["Klingler_et_al._BODIE_PIKA_DATA_DRYAD.xlsx"], readme_text)
        result["geometry"]["workbook_diagnostics"] = scan["diagnostics"]
        result["geometry"]["candidate_summaries"] = scan["candidate_summaries"]
        if len(scan["valid"]) != 1:
            result["status"] = CONTRACT["terminal_stop"]["status_if_complete_geometry_not_proven"]
            result["reason"] = f"required exactly one complete one-row-per-patch documented geometry table with {EXPECTED_N} unique patches; found {len(scan['valid'])}"
        else:
            chosen = scan["valid"][0]
            recs = sorted(chosen["records"], key=lambda x: x[0])
            ladder_rows, distinct = ladder(recs, chosen["mode"])
            registry_obj = [{"patch_id": r[0], "coord_1": r[1], "coord_2": r[2], "mode": chosen["mode"]} for r in recs]
            result["geometry"].update({
                "selected_sheet": chosen["sheet"],
                "selected_header_row": chosen["header_row"],
                "coordinate_mode": chosen["mode"],
                "patch_count": len(recs),
                "registry_fingerprint": canonical_sha256(registry_obj),
                "structural_ladder": ladder_rows,
                "distinct_positive_threshold_count": distinct
            })
            if distinct < MIN_SCALES:
                result["status"] = CONTRACT["terminal_stop"]["status_if_structural_scales_insufficient"]
                result["reason"] = f"only {distinct} distinct positive response-blind structural thresholds; require >= {MIN_SCALES}"
            else:
                result["status"] = "gate0_pass_response_independent_complete_geometry"
                result["reason"] = "complete 82-patch response-independent geometry and structural-scale diversity reproduced without census payload access"
        result["fingerprint"] = canonical_sha256({k: v for k, v in result.items() if k != "fingerprint"})
        (BUILD / "gate0_geometry.json").write_text(json.dumps(result, indent=2, sort_keys=True) + "\n")
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0
    except Exception as exc:
        result["status"] = "engineering_failure_pre_response"
        result["reason"] = f"{type(exc).__name__}: {exc}"
        result["fingerprint"] = canonical_sha256({k: v for k, v in result.items() if k != "fingerprint"})
        (BUILD / "gate0_geometry.json").write_text(json.dumps(result, indent=2, sort_keys=True) + "\n")
        print(json.dumps(result, indent=2, sort_keys=True))
        return 1


if __name__ == "__main__":
    sys.exit(main())
