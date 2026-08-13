from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from benchmarks.thalassia_genalex_stage3 import (
    GenAlExSchemaError,
    build_pairwise_response,
    complete_case_clone_correct,
    parse_genalex_workbook,
)


def _write_workbook(
    path: Path,
    *,
    population_ids=("P1", "P2"),
    locus_names=("L1", "L2"),
    rows=None,
    extra_sheet=False,
):
    rows = rows or [
        ("s1", "P1", (100, 100), (200, 202)),
        ("s2", "P1", (100, 102), (200, 202)),
        ("s3", "P2", (102, 102), (202, 202)),
        ("s4", "P2", (100, 102), (200, 200)),
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Genotypes"
    sheet["A1"] = len(locus_names)
    sheet["B1"] = len(rows)
    sheet["C1"] = len(population_ids)
    sizes = [sum(row[1] == population_id for row in rows) for population_id in population_ids]
    for offset, (population_id, size) in enumerate(zip(population_ids, sizes, strict=True), start=4):
        sheet.cell(row=1, column=offset).value = size
        sheet.cell(row=2, column=offset).value = population_id
    sheet["A3"] = "Sample"
    sheet["B3"] = "Pop"
    for locus_index, locus_name in enumerate(locus_names):
        sheet.cell(row=3, column=3 + 2 * locus_index).value = locus_name
    for row_index, (sample_id, population_id, *genotypes) in enumerate(rows, start=4):
        sheet.cell(row=row_index, column=1).value = sample_id
        sheet.cell(row=row_index, column=2).value = population_id
        for locus_index, genotype in enumerate(genotypes):
            a, b = genotype
            sheet.cell(row=row_index, column=3 + 2 * locus_index).value = a
            sheet.cell(row=row_index, column=4 + 2 * locus_index).value = b
    if extra_sheet:
        duplicate = workbook.copy_worksheet(sheet)
        duplicate.title = "AnotherValidSheet"
    workbook.save(path)


def _parse(path: Path):
    return parse_genalex_workbook(
        path,
        population_ids=("P1", "P2"),
        site_names={},
        expected_loci=2,
        expected_sha256=None,
    )


def test_standard_genalex_codominant_layout_is_parsed(tmp_path):
    path = tmp_path / "toy.xlsx"
    _write_workbook(path)
    parsed = _parse(path)
    assert parsed.sheet_name == "Genotypes"
    assert parsed.locus_order == ("L1", "L2")
    assert parsed.raw_sample_count == 4
    assert parsed.samples[0].genotypes == ((100, 100), (200, 202))
    assert parsed.samples[0].source_row == 4


def test_zero_or_blank_in_either_allele_marks_locus_missing_for_complete_case(tmp_path):
    path = tmp_path / "missing.xlsx"
    _write_workbook(
        path,
        rows=[
            ("s1", "P1", (100, 100), (0, 0)),
            ("s2", "P1", (100, 102), (200, 202)),
            ("s3", "P2", (102, 102), (202, 202)),
            ("s4", "P2", (100, 102), (200, 200)),
        ],
    )
    parsed = _parse(path)
    assert parsed.samples[0].genotypes[1] is None

    partial = tmp_path / "partial.xlsx"
    _write_workbook(
        partial,
        rows=[
            ("s1", "P1", (100, 100), (0, 202)),
            ("s2", "P1", (100, 102), (200, 202)),
            ("s3", "P2", (102, 102), (202, 202)),
            ("s4", "P2", (100, 102), (200, 200)),
        ],
    )
    parsed_partial = _parse(partial)
    assert parsed_partial.samples[0].genotypes[1] is None


def test_complete_case_filter_then_exact_within_population_clone_correction(tmp_path):
    path = tmp_path / "clones.xlsx"
    _write_workbook(
        path,
        rows=[
            ("z-clone", "P1", (100, 100), (200, 200)),
            ("a-clone", "P1", (100, 100), (200, 200)),
            ("incomplete", "P1", (100, 102), (0, 0)),
            ("unique", "P1", (102, 102), (202, 202)),
            ("p2a", "P2", (100, 102), (200, 202)),
            ("p2b", "P2", (102, 102), (202, 202)),
        ],
    )
    parsed = _parse(path)
    corrected, qc = complete_case_clone_correct(parsed, population_order=("P1", "P2"))
    assert [sample.sample_id for sample in corrected["P1"]] == ["a-clone", "unique"]
    assert qc == {
        "dropped_incomplete_samples": 1,
        "removed_exact_within_population_clonemates": 1,
    }


def test_clone_representative_uses_normalized_sample_id_then_source_row(tmp_path):
    path = tmp_path / "normalized-clones.xlsx"
    _write_workbook(
        path,
        rows=[
            ("A-01", "P1", (100, 100), (200, 200)),
            ("a01", "P1", (100, 100), (200, 200)),
            ("unique", "P1", (102, 102), (202, 202)),
            ("p2a", "P2", (100, 102), (200, 202)),
            ("p2b", "P2", (102, 102), (202, 202)),
        ],
    )
    parsed = _parse(path)
    corrected, _ = complete_case_clone_correct(parsed, population_order=("P1", "P2"))
    assert [sample.sample_id for sample in corrected["P1"]] == ["A-01", "unique"]


def test_population_blocks_must_match_declared_genalex_parameters(tmp_path):
    path = tmp_path / "wrong-block.xlsx"
    _write_workbook(
        path,
        rows=[
            ("s1", "P1", (100, 100), (200, 202)),
            ("s2", "P2", (102, 102), (202, 202)),
            ("s3", "P1", (100, 102), (200, 202)),
            ("s4", "P2", (100, 102), (200, 200)),
        ],
    )
    with pytest.raises(GenAlExSchemaError, match="population blocks"):
        _parse(path)


def test_exactly_one_candidate_genalex_sheet_is_required(tmp_path):
    path = tmp_path / "ambiguous.xlsx"
    _write_workbook(path, extra_sheet=True)
    with pytest.raises(GenAlExSchemaError, match="exactly one"):
        _parse(path)


def test_response_builder_rejects_negative_theta_for_frozen_linearized_domain(monkeypatch, tmp_path):
    path = tmp_path / "response.xlsx"
    _write_workbook(path)
    parsed = _parse(path)

    class Pair:
        population_a = "P1"
        population_b = "P2"
        theta = -0.05
        n_usable_loci = 2

    monkeypatch.setattr(
        "benchmarks.thalassia_genalex_stage3.all_pairwise_theta",
        lambda *args, **kwargs: (Pair(),),
    )
    with pytest.raises(GenAlExSchemaError, match="non_estimable_linearized_fst_domain"):
        build_pairwise_response(parsed, population_order=("P1", "P2"))
