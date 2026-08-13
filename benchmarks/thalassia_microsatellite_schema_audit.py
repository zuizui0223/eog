"""First authorized open of the frozen Thalassia microsatellite workbook.

This script validates only the GenAlEx schema, sample/population alignment, locus headers,
and missing-value representation. It deliberately does not identify multilocus genotypes,
calculate allele frequencies, or compute any genetic-distance/FST response.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_RAW_SHA256 = "aaaab9e302c9be8cf4e108d2aee5867c0b64ed70b6d3b81bad4d60168ee7f2f3"
EXPECTED_RAW_SIZE = 118400
EXPECTED_STAGE2_MANIFEST_SHA256 = "38f677f4d5f864d7d661e251c4e3970f2098ae2155a0125bb88045aa55fc442f"
EXPECTED_STAGE2_FINGERPRINT = "ef119675a596fe2044aca97b43efc618cfb7b00aee1dc3a1663ff5736c0a94ea"
EXPECTED_IDS = (
    "BIA", "TUA", "AMB", "KEN", "BIT", "PAL", "JEP", "PAR", "BAN",
    "NAT", "KUP", "MAT", "DRI", "PAD", "CK", "MI", "Ex2",
)
N_LOCI = 16
N_POPS = 17
COORDINATES = Path("benchmarks/frozen/thalassia_response_free/population_coordinates.csv")
STAGE2_MANIFEST = Path("benchmarks/frozen/thalassia_stage2_predictors/manifest.json")
STAGE3_CONTRACT = Path("docs/eog_v2_thalassia_stage3_microsatellite_contract.md")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(value: object) -> str:
    encoded = json.dumps(value, sort_keys=True, separators=(",", ":"), allow_nan=False).encode()
    return hashlib.sha256(encoded).hexdigest()


def _text(value: Any) -> str:
    return str(value if value is not None else "").strip()


def _norm(value: Any) -> str:
    return "".join(ch for ch in _text(value).lower() if ch.isalnum())


def _int_parameter(value: Any, label: str) -> int:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} is not numeric") from exc
    if not math.isfinite(number) or abs(number - round(number)) > 1e-9:
        raise ValueError(f"{label} is not a finite integer")
    return int(round(number))


def _allele_code(value: Any) -> tuple[str, int | None]:
    """Return (class, integer allele). Class is complete/missing/invalid."""
    if value is None or _text(value) == "":
        return "missing", None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "invalid", None
    if not math.isfinite(number) or abs(number - round(number)) > 1e-9:
        return "invalid", None
    integer = int(round(number))
    if integer == 0:
        return "missing", None
    if integer < 0:
        return "invalid", None
    return "complete", integer


def _frozen_aliases() -> tuple[dict[str, str], list[dict[str, str]]]:
    aliases: dict[str, str] = {}
    rows: list[dict[str, str]] = []
    with COORDINATES.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            code = raw["population_id"].strip()
            name = raw["site_name"].strip()
            rows.append({"code": code, "site_name": name})
            for token in (code, name):
                normalized = _norm(token)
                if not normalized:
                    raise ValueError("empty frozen population alias")
                previous = aliases.get(normalized)
                if previous is not None and previous != code:
                    raise ValueError(f"ambiguous frozen population alias: {token!r}")
                aliases[normalized] = code
    if tuple(row["code"] for row in rows) != EXPECTED_IDS:
        raise ValueError("frozen Thalassia coordinate population order drifted")
    return aliases, rows


def _map_population(value: Any, aliases: dict[str, str]) -> str:
    normalized = _norm(value)
    if not normalized or normalized not in aliases:
        raise ValueError(f"unmapped Thalassia population token: {_text(value)!r}")
    return aliases[normalized]


def audit(raw_path: Path) -> dict[str, object]:
    if not STAGE3_CONTRACT.exists():
        raise RuntimeError("Thalassia Stage-3 response contract is absent")
    if raw_path.stat().st_size != EXPECTED_RAW_SIZE or _sha256(raw_path) != EXPECTED_RAW_SHA256:
        raise ValueError("Thalassia microsatellite workbook identity mismatch")
    if _sha256(STAGE2_MANIFEST) != EXPECTED_STAGE2_MANIFEST_SHA256:
        raise ValueError("Thalassia frozen Stage-2 manifest byte identity drifted")
    stage2 = json.loads(STAGE2_MANIFEST.read_text(encoding="utf-8"))
    if stage2.get("fingerprint") != EXPECTED_STAGE2_FINGERPRINT:
        raise ValueError("Thalassia frozen Stage-2 manifest fingerprint drifted")
    if stage2.get("genetic_response_attached") is not False:
        raise ValueError("Thalassia Stage-2 predictors are no longer response-free")

    aliases, frozen_alias_rows = _frozen_aliases()
    # Normal worksheet mode caches cells once.  This changes no schema or response rule;
    # it only avoids the repeated XML rescans caused by random .cell() access on a
    # ReadOnlyWorksheet.
    workbook = load_workbook(raw_path, read_only=False, data_only=True)
    sheet_summaries: list[dict[str, object]] = []
    candidates: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        n_rows = int(sheet.max_row or 0)
        n_cols = int(sheet.max_column or 0)
        summary: dict[str, object] = {"sheet": sheet.title, "max_row": n_rows, "max_column": n_cols}
        sheet_summaries.append(summary)
        if n_rows < 4 or n_cols < 2 + 2 * N_LOCI:
            continue
        try:
            n_loci = _int_parameter(sheet.cell(1, 1).value, f"{sheet.title}!A1")
            n_samples = _int_parameter(sheet.cell(1, 2).value, f"{sheet.title}!B1")
            n_pops = _int_parameter(sheet.cell(1, 3).value, f"{sheet.title}!C1")
        except ValueError:
            continue
        if (n_loci, n_pops) != (N_LOCI, N_POPS) or n_samples <= 0:
            continue
        if n_cols < 2 + 2 * n_loci:
            continue

        pop_sizes = [
            _int_parameter(sheet.cell(1, 4 + index).value, f"{sheet.title}!pop_size_{index+1}")
            for index in range(n_pops)
        ]
        if any(value <= 0 for value in pop_sizes) or sum(pop_sizes) != n_samples:
            raise ValueError(f"{sheet.title}: declared population sizes do not sum to sample count")
        declared_pop_tokens = [_text(sheet.cell(2, 4 + index).value) for index in range(n_pops)]
        declared_pop_ids = [_map_population(token, aliases) for token in declared_pop_tokens]
        if len(set(declared_pop_ids)) != N_POPS or set(declared_pop_ids) != set(EXPECTED_IDS):
            raise ValueError(f"{sheet.title}: declared population labels do not bijectively map to the 17 frozen nodes")

        sample_header = _text(sheet.cell(3, 1).value)
        population_header = _text(sheet.cell(3, 2).value)
        locus_names: list[str] = []
        for locus_index in range(n_loci):
            first_col = 3 + 2 * locus_index
            second_col = first_col + 1
            locus_name = _text(sheet.cell(3, first_col).value)
            second_header = _text(sheet.cell(3, second_col).value)
            if not locus_name:
                raise ValueError(f"{sheet.title}: locus {locus_index+1} first allele column lacks a header")
            if second_header not in {"", locus_name}:
                raise ValueError(
                    f"{sheet.title}: locus {locus_index+1} second allele header is not blank/the first header"
                )
            locus_names.append(locus_name)
        if len(set(locus_names)) != n_loci:
            raise ValueError(f"{sheet.title}: locus headers are not unique")

        data_rows: list[dict[str, object]] = []
        sample_ids: list[str] = []
        mapped_population_ids: list[str] = []
        missing_allele_cells = 0
        partial_missing_locus_pairs = 0
        fully_missing_locus_pairs = 0
        invalid_allele_cells = 0
        for offset in range(n_samples):
            row_number = 4 + offset
            sample_id = _text(sheet.cell(row_number, 1).value)
            if not sample_id:
                raise ValueError(f"{sheet.title}: empty sample ID at data row {row_number}")
            population_token = sheet.cell(row_number, 2).value
            population_id = _map_population(population_token, aliases)
            sample_ids.append(sample_id)
            mapped_population_ids.append(population_id)
            row_missing = 0
            row_partial = 0
            row_fully_missing = 0
            for locus_index in range(n_loci):
                first_col = 3 + 2 * locus_index
                left_class, _ = _allele_code(sheet.cell(row_number, first_col).value)
                right_class, _ = _allele_code(sheet.cell(row_number, first_col + 1).value)
                invalid_allele_cells += int(left_class == "invalid") + int(right_class == "invalid")
                row_missing += int(left_class == "missing") + int(right_class == "missing")
                if (left_class == "missing") ^ (right_class == "missing"):
                    row_partial += 1
                if left_class == "missing" and right_class == "missing":
                    row_fully_missing += 1
            missing_allele_cells += row_missing
            partial_missing_locus_pairs += row_partial
            fully_missing_locus_pairs += row_fully_missing
            data_rows.append(
                {
                    "row": row_number,
                    "sample_id": sample_id,
                    "population_token": _text(population_token),
                    "population_id": population_id,
                    "missing_allele_cells": row_missing,
                    "partial_missing_locus_pairs": row_partial,
                    "fully_missing_locus_pairs": row_fully_missing,
                }
            )

        if invalid_allele_cells:
            raise ValueError(f"{sheet.title}: encountered {invalid_allele_cells} invalid allele cells")
        if len(set(sample_ids)) != n_samples:
            raise ValueError(f"{sheet.title}: sample IDs are not unique")
        mapped_counts = {population_id: mapped_population_ids.count(population_id) for population_id in EXPECTED_IDS}
        if any(value <= 0 for value in mapped_counts.values()):
            raise ValueError(f"{sheet.title}: not all 17 frozen populations appear in data rows")

        cursor = 0
        block_checks: list[dict[str, object]] = []
        for declared_id, declared_size in zip(declared_pop_ids, pop_sizes):
            block = mapped_population_ids[cursor:cursor + declared_size]
            block_ok = len(block) == declared_size and set(block) == {declared_id}
            block_checks.append(
                {"population_id": declared_id, "declared_size": declared_size, "contiguous_block_ok": block_ok}
            )
            cursor += declared_size
        if cursor != n_samples or not all(bool(item["contiguous_block_ok"]) for item in block_checks):
            raise ValueError(f"{sheet.title}: population data rows do not match declared contiguous GenAlEx blocks")

        candidates.append(
            {
                "sheet": sheet.title,
                "n_loci": n_loci,
                "n_samples": n_samples,
                "n_populations": n_pops,
                "sample_header": sample_header,
                "population_header": population_header,
                "locus_names": locus_names,
                "declared_population_tokens": declared_pop_tokens,
                "declared_population_ids": declared_pop_ids,
                "population_sizes": pop_sizes,
                "mapped_population_counts": mapped_counts,
                "block_checks": block_checks,
                "missing_allele_cells": missing_allele_cells,
                "partial_missing_locus_pairs": partial_missing_locus_pairs,
                "fully_missing_locus_pairs": fully_missing_locus_pairs,
                "data_rows_schema": data_rows,
            }
        )

    if len(candidates) != 1:
        raise RuntimeError(
            f"non_estimable_microsatellite_schema_or_population_alignment: candidate_blocks={len(candidates)}"
        )

    admitted = candidates[0]
    schema_payload = {
        "sheet": admitted["sheet"],
        "n_loci": admitted["n_loci"],
        "n_samples": admitted["n_samples"],
        "n_populations": admitted["n_populations"],
        "locus_names": admitted["locus_names"],
        "declared_population_ids": admitted["declared_population_ids"],
        "population_sizes": admitted["population_sizes"],
        "sample_ids": [row["sample_id"] for row in admitted["data_rows_schema"]],
        "mapped_population_ids": [row["population_id"] for row in admitted["data_rows_schema"]],
    }
    result: dict[str, object] = {
        "status": "thalassia-microsatellite-schema-admitted",
        "raw_file_name": raw_path.name,
        "raw_size_bytes": raw_path.stat().st_size,
        "raw_sha256": _sha256(raw_path),
        "stage2_manifest_sha256": EXPECTED_STAGE2_MANIFEST_SHA256,
        "stage2_manifest_fingerprint": EXPECTED_STAGE2_FINGERPRINT,
        "stage3_contract_sha256": _sha256(STAGE3_CONTRACT),
        "sheet_summaries": sheet_summaries,
        "admitted_block": admitted,
        "frozen_population_aliases": frozen_alias_rows,
        "schema_fingerprint": _canonical_sha256(schema_payload),
        "genetic_workbook_contents_accessed_for_schema_only": True,
        "allele_frequencies_computed": False,
        "multilocus_genotypes_computed": False,
        "genetic_response_computed": False,
        "claim_boundary": (
            "The exact workbook was opened for GenAlEx schema/alignment validation only. No allele-frequency, "
            "MLG, FST, cluster, assignment, migration or genetic-distance response was computed."
        ),
    }
    result["fingerprint"] = _canonical_sha256(result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = audit(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2, sort_keys=True, allow_nan=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, sort_keys=True, allow_nan=False))


if __name__ == "__main__":
    main()
