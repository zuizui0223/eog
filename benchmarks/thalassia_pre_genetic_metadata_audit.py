"""Audit only the response-free Thalassia GPS workbook.

The genetic workbook path is intentionally not accepted.  Opaque genetic provenance is
passed as size/SHA strings computed by the workflow before this script is invoked.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MISSING = {"", "na", "n/a", "nan", "none", "null", ".", "-"}
SITE_TOKENS = ("site", "location", "locality", "station", "code", "id")
LAT_TOKENS = ("lat", "latitude")
LON_TOKENS = ("lon", "long", "longitude")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(value: object) -> str:
    return hashlib.sha256(json.dumps(value, sort_keys=True, separators=(",", ":"), allow_nan=False).encode()).hexdigest()


def _text(value: Any) -> str:
    return str(value if value is not None else "").strip()


def _norm(value: Any) -> str:
    return "".join(ch for ch in _text(value).lower() if ch.isalnum())


def _finite_number(value: Any, bound: float) -> float | None:
    text = _text(value)
    if text.lower() in MISSING:
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed) or not -bound <= parsed <= bound:
        return None
    return parsed


def audit(gps_path: Path, genetic_size: int, genetic_sha256: str) -> dict[str, object]:
    gps_sha = _sha256(gps_path)
    workbook = load_workbook(gps_path, read_only=True, data_only=True)
    sheets: list[dict[str, object]] = []
    coordinate_candidates: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        matrix = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        nonempty = [row for row in matrix if any(_text(value) for value in row)]
        if not nonempty:
            sheets.append({"name": sheet.title, "n_nonempty_rows": 0, "n_columns": 0, "header": []})
            continue
        width = max(len(row) for row in nonempty)
        header = list(nonempty[0]) + [None] * max(0, width - len(nonempty[0]))
        header_norm = [_norm(value) for value in header]
        site_cols = [i for i, value in enumerate(header_norm) if any(token in value for token in SITE_TOKENS)]
        lat_cols = [i for i, value in enumerate(header_norm) if any(token in value for token in LAT_TOKENS)]
        lon_cols = [i for i, value in enumerate(header_norm) if any(token in value for token in LON_TOKENS)]
        sheets.append(
            {
                "name": sheet.title,
                "n_nonempty_rows": len(nonempty),
                "n_columns": width,
                "header": [_text(value) for value in header],
                "site_candidate_columns": site_cols,
                "latitude_candidate_columns": lat_cols,
                "longitude_candidate_columns": lon_cols,
            }
        )
        for site_col in site_cols:
            for lat_col in lat_cols:
                for lon_col in lon_cols:
                    rows: list[dict[str, object]] = []
                    for row_index, raw in enumerate(nonempty[1:], start=2):
                        padded = list(raw) + [None] * max(0, width - len(raw))
                        site_id = _text(padded[site_col])
                        latitude = _finite_number(padded[lat_col], 90.0)
                        longitude = _finite_number(padded[lon_col], 180.0)
                        if site_id and latitude is not None and longitude is not None:
                            rows.append(
                                {
                                    "row": row_index,
                                    "site_id": site_id,
                                    "latitude": latitude,
                                    "longitude": longitude,
                                }
                            )
                    if rows:
                        coordinate_candidates.append(
                            {
                                "sheet": sheet.title,
                                "site_column": site_col,
                                "latitude_column": lat_col,
                                "longitude_column": lon_col,
                                "n_complete_rows": len(rows),
                                "n_unique_site_ids": len({str(row['site_id']) for row in rows}),
                                "rows": rows,
                            }
                        )

    result: dict[str, object] = {
        "status": "thalassia-response-free-gps-schema-audit",
        "gps_file_name": gps_path.name,
        "gps_size_bytes": gps_path.stat().st_size,
        "gps_sha256": gps_sha,
        "genetic_file_size_bytes": int(genetic_size),
        "genetic_file_sha256": genetic_sha256,
        "sheets": sheets,
        "coordinate_candidates": coordinate_candidates,
        "genetic_workbook_contents_accessed": False,
        "genetic_workbook_container_inspected": False,
        "genetic_response_computed": False,
        "claim_boundary": (
            "Only GPS workbook cells were opened. The microsatellite workbook was represented to this script solely by "
            "opaque file size/SHA-256 strings and was not opened as a workbook or archive."
        ),
    }
    result["fingerprint"] = _canonical_sha256(result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--gps", type=Path, required=True)
    parser.add_argument("--genetic-size", type=int, required=True)
    parser.add_argument("--genetic-sha256", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = audit(args.gps, args.genetic_size, args.genetic_sha256)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2, sort_keys=True, allow_nan=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, sort_keys=True, allow_nan=False))


if __name__ == "__main__":
    main()
