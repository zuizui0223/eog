"""Frozen Stage-3 GenAlEx parser and clone-corrected FST response builder for Thalassia.

The biological response rules in this module follow the first Stage-3 contract frozen before
the schema-only workbook access.  The real pairwise FST response must only be computed after
the schema artifact and this harmonized implementation are frozen.
"""
from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import hashlib
import json
import math
from pathlib import Path
import re
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook

from benchmarks.zhoushan_weir_cockerham import all_pairwise_theta

EXPECTED_GENETIC_SHA256 = "aaaab9e302c9be8cf4e108d2aee5867c0b64ed70b6d3b81bad4d60168ee7f2f3"
FROZEN_POPULATION_IDS = (
    "BIA", "TUA", "AMB", "KEN", "BIT", "PAL", "JEP", "PAR", "BAN",
    "NAT", "KUP", "MAT", "DRI", "PAD", "CK", "MI", "Ex2",
)
FROZEN_SITE_NAMES = {
    "BIA": "Biak",
    "TUA": "Tual",
    "AMB": "Ambon",
    "KEN": "Kendari",
    "BIT": "Bitung/Kema",
    "PAL": "Palu",
    "JEP": "Jepara",
    "PAR": "Pari",
    "BAN": "Bangka",
    "NAT": "Natuna",
    "KUP": "Kupang",
    "MAT": "Lombok",
    "DRI": "Drini",
    "PAD": "Padang",
    "CK": "Cocos Keeling",
    "MI": "Mermaid Island",
    "Ex2": "Muiron Island north",
}
EXPECTED_LOCI = 16
MISSING_ALLELE = 0


class GenAlExSchemaError(ValueError):
    """The workbook/response does not satisfy the frozen Stage-3 contract."""


@dataclass(frozen=True)
class SampleRecord:
    sample_id: str
    population_id: str
    genotypes: tuple[tuple[int, int] | None, ...]
    source_row: int


@dataclass(frozen=True)
class ParsedGenAlEx:
    sheet_name: str
    locus_order: tuple[str, ...]
    samples: tuple[SampleRecord, ...]
    declared_population_sizes: tuple[tuple[str, int], ...]
    raw_sample_count: int


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(value: object) -> str:
    payload = json.dumps(value, sort_keys=True, separators=(",", ":"), allow_nan=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _text(value: object) -> str:
    return str(value if value is not None else "").strip()


def _norm(value: object) -> str:
    return re.sub(r"[^0-9a-z]+", "", _text(value).casefold())


def _strict_int(value: object, *, label: str, minimum: int = 0) -> int:
    if isinstance(value, bool) or value is None or _text(value) == "":
        raise GenAlExSchemaError(f"{label} must be an integer")
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise GenAlExSchemaError(f"{label} must be numeric") from exc
    if not math.isfinite(numeric) or numeric != math.floor(numeric):
        raise GenAlExSchemaError(f"{label} must be a finite integer")
    result = int(numeric)
    if result < minimum:
        raise GenAlExSchemaError(f"{label} must be >= {minimum}")
    return result


def _parse_allele(value: object, *, label: str) -> int:
    # The first frozen contract treats both blank and GenAlEx 0 as missing.  A one-sided
    # missing pair makes the whole locus missing for complete-case filtering; it is not
    # imputed and is not a schema rescue.
    if value is None or _text(value) == "":
        return MISSING_ALLELE
    return _strict_int(value, label=label, minimum=0)


def population_aliases(
    population_ids: Sequence[str],
    *,
    site_names: Mapping[str, str] | None = None,
) -> dict[str, str]:
    ids = tuple(str(value) for value in population_ids)
    if not ids or len(ids) != len(set(ids)):
        raise ValueError("population_ids must contain unique IDs")
    aliases: dict[str, str] = {}
    for population_id in ids:
        for value in (population_id, (site_names or {}).get(population_id)):
            if value is None:
                continue
            key = _norm(value)
            if not key:
                raise ValueError("population aliases cannot normalize to empty strings")
            previous = aliases.get(key)
            if previous is not None and previous != population_id:
                raise ValueError(f"ambiguous frozen population alias: {value!r}")
            aliases[key] = population_id
    return aliases


def _map_population(value: object, aliases: Mapping[str, str], *, label: str) -> str:
    key = _norm(value)
    if not key or key not in aliases:
        raise GenAlExSchemaError(f"{label} does not match a frozen population ID/site label")
    return aliases[key]


def _parse_candidate_sheet(
    sheet,
    *,
    population_ids: Sequence[str],
    aliases: Mapping[str, str],
    expected_loci: int,
) -> ParsedGenAlEx | None:
    try:
        n_loci = _strict_int(sheet["A1"].value, label=f"{sheet.title}!A1", minimum=1)
        n_populations = _strict_int(sheet["C1"].value, label=f"{sheet.title}!C1", minimum=1)
    except GenAlExSchemaError:
        return None
    if n_loci != expected_loci or n_populations != len(population_ids):
        return None

    n_samples = _strict_int(sheet["B1"].value, label=f"{sheet.title}!B1", minimum=1)
    population_row_labels: list[str] = []
    declared_sizes: list[int] = []
    for offset in range(len(population_ids)):
        column = 4 + offset
        raw_label = sheet.cell(row=2, column=column).value
        population_row_labels.append(
            _map_population(raw_label, aliases, label=f"{sheet.title}!{sheet.cell(row=2, column=column).coordinate}")
        )
        declared_sizes.append(
            _strict_int(
                sheet.cell(row=1, column=column).value,
                label=f"{sheet.title}!{sheet.cell(row=1, column=column).coordinate}",
                minimum=1,
            )
        )
    if len(set(population_row_labels)) != len(population_ids) or set(population_row_labels) != set(population_ids):
        raise GenAlExSchemaError("GenAlEx row-2 population labels must match the frozen population set exactly")
    if sum(declared_sizes) != n_samples:
        raise GenAlExSchemaError("GenAlEx population sizes must sum to B1 sample count")

    locus_order: list[str] = []
    for locus_index in range(expected_loci):
        column = 3 + 2 * locus_index
        locus_name = _text(sheet.cell(row=3, column=column).value)
        if not locus_name:
            raise GenAlExSchemaError(f"missing locus label at {sheet.cell(row=3, column=column).coordinate}")
        locus_order.append(locus_name)
    if len(set(locus_order)) != expected_loci:
        raise GenAlExSchemaError("locus labels must be unique")

    samples: list[SampleRecord] = []
    sample_ids: set[str] = set()
    data_start = 4
    data_end = data_start + n_samples - 1
    genotype_end_column = 2 + 2 * expected_loci

    for row_index in range(data_start, data_end + 1):
        sample_id = _text(sheet.cell(row=row_index, column=1).value)
        if not sample_id or sample_id in sample_ids:
            raise GenAlExSchemaError(f"sample IDs must be non-empty and unique; row {row_index}")
        sample_ids.add(sample_id)
        population_id = _map_population(
            sheet.cell(row=row_index, column=2).value,
            aliases,
            label=f"{sheet.title}!B{row_index}",
        )
        genotypes: list[tuple[int, int] | None] = []
        for locus_index, locus_name in enumerate(locus_order):
            column = 3 + 2 * locus_index
            a = _parse_allele(
                sheet.cell(row=row_index, column=column).value,
                label=f"{sheet.title}!{sheet.cell(row=row_index, column=column).coordinate}",
            )
            b = _parse_allele(
                sheet.cell(row=row_index, column=column + 1).value,
                label=f"{sheet.title}!{sheet.cell(row=row_index, column=column + 1).coordinate}",
            )
            genotypes.append(None if (a == MISSING_ALLELE or b == MISSING_ALLELE) else tuple(sorted((a, b))))
        samples.append(SampleRecord(sample_id, population_id, tuple(genotypes), row_index))

    cursor = 0
    for population_id, declared_size in zip(population_row_labels, declared_sizes, strict=True):
        observed = [sample.population_id for sample in samples[cursor: cursor + declared_size]]
        if observed != [population_id] * declared_size:
            raise GenAlExSchemaError("sample population blocks do not match frozen GenAlEx row-1/row-2 parameters")
        cursor += declared_size

    for row_index in range(data_end + 1, sheet.max_row + 1):
        if any(_text(sheet.cell(row=row_index, column=column).value) for column in range(1, genotype_end_column + 1)):
            raise GenAlExSchemaError("unexpected non-empty cells below the declared GenAlEx sample block")

    declared = tuple(zip(population_row_labels, declared_sizes, strict=True))
    return ParsedGenAlEx(
        sheet_name=sheet.title,
        locus_order=tuple(locus_order),
        samples=tuple(samples),
        declared_population_sizes=declared,
        raw_sample_count=n_samples,
    )


def parse_genalex_workbook(
    path: str | Path,
    *,
    population_ids: Sequence[str] = FROZEN_POPULATION_IDS,
    site_names: Mapping[str, str] | None = FROZEN_SITE_NAMES,
    expected_loci: int = EXPECTED_LOCI,
    expected_sha256: str | None = EXPECTED_GENETIC_SHA256,
) -> ParsedGenAlEx:
    source = Path(path)
    if source.suffix.lower() != ".xlsx":
        raise GenAlExSchemaError("released Thalassia microsatellite source must be .xlsx")
    if expected_sha256 is not None:
        actual = _sha256(source)
        if actual != expected_sha256:
            raise GenAlExSchemaError(f"microsatellite SHA-256 mismatch: {actual}")

    aliases = population_aliases(population_ids, site_names=site_names)
    # Cached worksheet cells avoid repeated XML rescans during the one-time exact response.
    workbook = load_workbook(source, read_only=False, data_only=True)
    matches: list[ParsedGenAlEx] = []
    for sheet in workbook.worksheets:
        parsed = _parse_candidate_sheet(
            sheet,
            population_ids=population_ids,
            aliases=aliases,
            expected_loci=expected_loci,
        )
        if parsed is not None:
            matches.append(parsed)
    if len(matches) != 1:
        raise GenAlExSchemaError(
            f"expected exactly one standard GenAlEx codominant sheet; found {len(matches)}"
        )
    return matches[0]


def _complete_mlg(sample: SampleRecord) -> tuple[tuple[int, int], ...] | None:
    if any(genotype is None for genotype in sample.genotypes):
        return None
    return tuple(genotype for genotype in sample.genotypes if genotype is not None)


def _representative_key(sample: SampleRecord) -> tuple[str, int]:
    return _norm(sample.sample_id), sample.source_row


def complete_case_clone_correct(
    parsed: ParsedGenAlEx,
    *,
    population_order: Sequence[str],
) -> tuple[dict[str, tuple[SampleRecord, ...]], dict[str, int]]:
    """Drop incomplete samples, then retain one exact within-population MLG representative."""
    by_population: dict[str, list[SampleRecord]] = {str(pid): [] for pid in population_order}
    if set(by_population) != {sample.population_id for sample in parsed.samples}:
        raise GenAlExSchemaError("parsed population set does not match the frozen population set")

    dropped_incomplete = 0
    duplicate_clonemates = 0
    for population_id in by_population:
        complete = [
            sample for sample in parsed.samples
            if sample.population_id == population_id and _complete_mlg(sample) is not None
        ]
        dropped_incomplete += sum(
            1 for sample in parsed.samples
            if sample.population_id == population_id and _complete_mlg(sample) is None
        )
        groups: dict[tuple[tuple[int, int], ...], list[SampleRecord]] = {}
        for sample in complete:
            key = _complete_mlg(sample)
            assert key is not None
            groups.setdefault(key, []).append(sample)
        retained: list[SampleRecord] = []
        for key in sorted(groups, key=repr):
            group = sorted(groups[key], key=_representative_key)
            retained.append(group[0])
            duplicate_clonemates += len(group) - 1
        retained.sort(key=_representative_key)
        if len(retained) < 2:
            raise GenAlExSchemaError(
                f"population {population_id} has fewer than two complete unique MLGs after frozen filtering"
            )
        by_population[population_id] = retained

    return (
        {population_id: tuple(records) for population_id, records in by_population.items()},
        {
            "dropped_incomplete_samples": dropped_incomplete,
            "removed_exact_within_population_clonemates": duplicate_clonemates,
        },
    )


def build_pairwise_response(
    parsed: ParsedGenAlEx,
    *,
    population_order: Sequence[str] = FROZEN_POPULATION_IDS,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    corrected, qc = complete_case_clone_correct(parsed, population_order=population_order)
    populations: dict[str, dict[str, list[tuple[int, int]]]] = {}
    for population_id in population_order:
        records = corrected[str(population_id)]
        populations[str(population_id)] = {
            locus: [record.genotypes[index] for record in records]  # type: ignore[list-item]
            for index, locus in enumerate(parsed.locus_order)
        }

    pairwise = all_pairwise_theta(
        populations,
        population_order=tuple(str(value) for value in population_order),
        locus_order=parsed.locus_order,
    )
    rows: list[dict[str, object]] = []
    for result in pairwise:
        theta = float(result.theta)
        # The first frozen Stage-3 contract and the already frozen generic linearized_fst
        # implementation require the full primary response domain 0 <= theta < 1.
        if not math.isfinite(theta) or theta < 0.0 or theta >= 1.0:
            raise GenAlExSchemaError(
                "non_estimable_linearized_fst_domain: "
                f"{result.population_a}-{result.population_b} theta={theta}"
            )
        linearized = theta / (1.0 - theta)
        if not math.isfinite(linearized):
            raise GenAlExSchemaError("non_estimable_linearized_fst_domain: non-finite linearized FST")
        rows.append(
            {
                "population_a": result.population_a,
                "population_b": result.population_b,
                "fst_theta": theta,
                "linearized_fst": linearized,
                "n_usable_loci": result.n_usable_loci,
            }
        )

    expected_pairs = len(population_order) * (len(population_order) - 1) // 2
    if len(rows) != expected_pairs:
        raise RuntimeError("pair-count drift")

    summary: dict[str, object] = {
        "sheet_name": parsed.sheet_name,
        "raw_sample_count": parsed.raw_sample_count,
        "n_loci": len(parsed.locus_order),
        "locus_order": list(parsed.locus_order),
        "population_order": [str(value) for value in population_order],
        "n_pairs": len(rows),
        "retained_unique_complete_mlg_by_population": {
            population_id: len(corrected[population_id]) for population_id in population_order
        },
        **qc,
        "clone_rule": (
            f"within-population exact complete {len(parsed.locus_order)}-locus MLG; "
            "retain smallest normalized sample ID, then source row"
        ),
        "missing_rule": (
            "GenAlEx 0 or blank in either allele makes that locus missing; complete-case samples only "
            "before clone correction; no imputation"
        ),
        "fst_estimator": "multilocus Weir-Cockerham (1984) theta; no clipping",
        "response_transform": "theta/(1-theta); primary domain requires 0<=theta<1",
    }
    summary["fingerprint"] = _canonical_sha256({"summary": summary, "pairs": rows})
    return rows, summary


def write_response_csv(rows: Iterable[Mapping[str, object]], path: Path) -> None:
    materialized = list(rows)
    fieldnames = ["population_a", "population_b", "fst_theta", "linearized_fst", "n_usable_loci"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(materialized)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--response-csv", type=Path, required=True)
    parser.add_argument("--summary-json", type=Path, required=True)
    args = parser.parse_args()

    parsed = parse_genalex_workbook(args.input)
    rows, summary = build_pairwise_response(parsed)
    write_response_csv(rows, args.response_csv)
    args.summary_json.parent.mkdir(parents=True, exist_ok=True)
    args.summary_json.write_text(
        json.dumps(summary, indent=2, sort_keys=True, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, sort_keys=True, allow_nan=False))


if __name__ == "__main__":
    main()
