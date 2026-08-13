"""One-time Stage-3B Thalassia genetic validation under the frozen v2.1 contract.

The response-free node universe, conventional references, EOG predictors, clone rule,
FST estimator, response transform, nested selector, bootstrap and GO rule were frozen
before any genetic response was computed.  Stage 3A opened the workbook for schema-only
inspection; this module is the first code authorized to construct MLGs or pairwise FST.
"""
from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass
import hashlib
import json
import math
from pathlib import Path
import re
from typing import Mapping, Sequence

import numpy as np

from benchmarks.zhoushan_weir_cockerham import all_pairwise_theta
from eog.genetic_reference_selection import evaluate_nested_genetic_reference_selection
from eog.genetic_validation import GeneticValidationConfig, _fit_ridge, _predict_ridge


EXPECTED_RAW_SHA256 = "aaaab9e302c9be8cf4e108d2aee5867c0b64ed70b6d3b81bad4d60168ee7f2f3"
EXPECTED_RAW_SIZE = 118400
EXPECTED_STAGE2_PREDICTORS_SHA256 = "05c3a2e65029d44532e55436bc97c33482f96fb2d1951d79a0b50c776b02b7e1"
EXPECTED_STAGE2_FINGERPRINT = "ef119675a596fe2044aca97b43efc618cfb7b00aee1dc3a1663ff5736c0a94ea"
EXPECTED_STAGE3A_SCHEMA_SHA256 = "2adacdadb9bd10488e4514dce583b57156f97e964037c5683592b83524038f2b"
EXPECTED_STAGE3A_SCHEMA_FINGERPRINT = "a0b2c6bb0755f1d09f9aa88fbe4bfa645c5a4c8f011907fe4ceb93f64ff674de"
EXPECTED_STAGE3A_AUDIT_FINGERPRINT = "3734a91c72f5f3fc7732e490216b9218b52f79a2506453039a8b2d72a01c03f9"

POPULATION_ORDER = (
    "BIA", "TUA", "AMB", "KEN", "BIT", "PAL", "JEP", "PAR", "BAN",
    "NAT", "KUP", "MAT", "DRI", "PAD", "CK", "MI", "Ex2",
)
POPULATION_SIZES = (48, 48, 47, 47, 48, 48, 46, 48, 48, 47, 48, 47, 48, 47, 47, 48, 46)
LOCUS_ORDER = (
    "Thh5", "Thh34", "Thh15", "TH66", "TH37", "TH73", "TH43", "Thh8",
    "TH34", "Thh41", "TH52", "TH07", "Thh29", "Thh1", "Thh36", "Thh3",
)
N_SAMPLES = 806
RIDGE_PENALTY = 1.0
BOOTSTRAP_SEED = 20260813
BOOTSTRAP_RESAMPLES = 10_000
CANDIDATE_IDS = ("gabriel_current_flow", "gabriel_shortest_path", "geographic")

STAGE2_PREDICTORS = Path("benchmarks/frozen/thalassia_stage2_predictors/predictors.csv")
STAGE2_MANIFEST = Path("benchmarks/frozen/thalassia_stage2_predictors/manifest.json")
STAGE3A_SCHEMA = Path("benchmarks/frozen/thalassia_stage3_schema/schema_audit.json")


class NonEstimable(RuntimeError):
    def __init__(self, status: str, detail: str):
        super().__init__(detail)
        self.status = status
        self.detail = detail


@dataclass(frozen=True)
class SampleGenotype:
    source_row: int
    sample_id: str
    population_id: str
    genotypes: tuple[tuple[int, int] | None, ...]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(value: object) -> str:
    payload = json.dumps(value, sort_keys=True, separators=(",", ":"), allow_nan=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _text(value: object) -> str:
    return str(value if value is not None else "").strip()


def _norm_sample_id(value: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", value.casefold())


def _allele(value: object, *, label: str) -> int | None:
    if value is None or _text(value) == "":
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"{label} non-numeric") from exc
    if not math.isfinite(numeric) or numeric != math.floor(numeric):
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"{label} non-integer")
    integer = int(numeric)
    if integer < 0:
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"{label} negative")
    return None if integer == 0 else integer


def validate_archived_inputs() -> dict[str, object]:
    if not STAGE3A_SCHEMA.exists():
        raise NonEstimable("non_estimable_stage3a_not_archived", "frozen Stage-3A schema artifact is absent")
    if _sha256(STAGE3A_SCHEMA) != EXPECTED_STAGE3A_SCHEMA_SHA256:
        raise NonEstimable("non_estimable_stage3a_archive_drift", "Stage-3A schema bytes drifted")
    schema = json.loads(STAGE3A_SCHEMA.read_text(encoding="utf-8"))
    if schema.get("status") != "thalassia-microsatellite-schema-admitted":
        raise NonEstimable("non_estimable_stage3a_archive_drift", "Stage-3A schema is not admitted")
    if schema.get("schema_fingerprint") != EXPECTED_STAGE3A_SCHEMA_FINGERPRINT:
        raise NonEstimable("non_estimable_stage3a_archive_drift", "Stage-3A schema fingerprint drifted")
    if schema.get("fingerprint") != EXPECTED_STAGE3A_AUDIT_FINGERPRINT:
        raise NonEstimable("non_estimable_stage3a_archive_drift", "Stage-3A audit fingerprint drifted")
    if schema.get("genetic_response_computed") is not False or schema.get("multilocus_genotypes_computed") is not False:
        raise NonEstimable("non_estimable_stage3a_archive_drift", "Stage-3A response boundary was crossed")

    if _sha256(STAGE2_PREDICTORS) != EXPECTED_STAGE2_PREDICTORS_SHA256:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "Stage-2 predictor bytes drifted")
    manifest = json.loads(STAGE2_MANIFEST.read_text(encoding="utf-8"))
    if manifest.get("fingerprint") != EXPECTED_STAGE2_FINGERPRINT:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "Stage-2 predictor fingerprint drifted")
    if tuple(manifest.get("candidate_ids") or ()) != CANDIDATE_IDS:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "candidate family/order drifted")
    if manifest.get("genetic_response_attached") is not False:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "Stage-2 bundle is no longer response-free")
    return {"stage2": manifest, "stage3a": schema}


def parse_genotypes(raw_path: Path) -> tuple[SampleGenotype, ...]:
    if raw_path.stat().st_size != EXPECTED_RAW_SIZE or _sha256(raw_path) != EXPECTED_RAW_SHA256:
        raise NonEstimable("non_estimable_raw_identity_mismatch", "Thalassia microsatellite workbook identity mismatch")

    # Imported only for the one-time workbook execution; package CI does not need openpyxl.
    from openpyxl import load_workbook

    workbook = load_workbook(raw_path, read_only=False, data_only=True)
    if workbook.sheetnames != ["Genalex"]:
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"sheet drift: {workbook.sheetnames!r}")
    sheet = workbook["Genalex"]
    if (sheet.max_row, sheet.max_column) != (809, 34):
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", "worksheet dimensions drifted")
    if (int(sheet["A1"].value), int(sheet["B1"].value), int(sheet["C1"].value)) != (16, N_SAMPLES, 17):
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", "GenAlEx parameters drifted")
    declared_sizes = tuple(int(sheet.cell(1, 4 + i).value) for i in range(17))
    declared_ids = tuple(_text(sheet.cell(2, 4 + i).value) for i in range(17))
    if declared_sizes != POPULATION_SIZES or declared_ids != POPULATION_ORDER:
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", "population parameters drifted")
    loci = tuple(_text(sheet.cell(3, 3 + 2 * i).value) for i in range(16))
    if loci != LOCUS_ORDER:
        raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", "locus order drifted")

    samples: list[SampleGenotype] = []
    sample_ids: set[str] = set()
    for offset in range(N_SAMPLES):
        row = 4 + offset
        sample_id = _text(sheet.cell(row, 1).value)
        population_id = _text(sheet.cell(row, 2).value)
        if not sample_id or sample_id in sample_ids:
            raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"sample ID drift at row {row}")
        sample_ids.add(sample_id)
        if population_id not in POPULATION_ORDER:
            raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"population drift at row {row}")
        genotypes: list[tuple[int, int] | None] = []
        for locus_index, locus in enumerate(LOCUS_ORDER):
            first_col = 3 + 2 * locus_index
            left = _allele(sheet.cell(row, first_col).value, label=f"row {row} {locus} allele1")
            right = _allele(sheet.cell(row, first_col + 1).value, label=f"row {row} {locus} allele2")
            # Frozen contract: if either allele is missing, the whole locus is missing.
            genotypes.append(None if left is None or right is None else tuple(sorted((left, right))))
        samples.append(SampleGenotype(row, sample_id, population_id, tuple(genotypes)))

    cursor = 0
    for population_id, declared_size in zip(POPULATION_ORDER, POPULATION_SIZES, strict=True):
        block = samples[cursor : cursor + declared_size]
        if [sample.population_id for sample in block] != [population_id] * declared_size:
            raise NonEstimable("non_estimable_microsatellite_schema_or_population_alignment", f"population block drift: {population_id}")
        cursor += declared_size
    return tuple(samples)


def complete_samples(samples: Sequence[SampleGenotype]) -> tuple[SampleGenotype, ...]:
    return tuple(sample for sample in samples if all(genotype is not None for genotype in sample.genotypes))


def clone_correct(
    samples: Sequence[SampleGenotype],
    *,
    population_order: Sequence[str] = POPULATION_ORDER,
) -> tuple[dict[str, tuple[SampleGenotype, ...]], dict[str, object]]:
    complete = complete_samples(samples)
    retained: dict[str, tuple[SampleGenotype, ...]] = {}
    duplicate_counts: dict[str, int] = {}
    raw_complete_counts: dict[str, int] = {}
    for population_id in population_order:
        population_samples = [sample for sample in complete if sample.population_id == population_id]
        raw_complete_counts[population_id] = len(population_samples)
        groups: dict[tuple[tuple[int, int], ...], list[SampleGenotype]] = {}
        for sample in population_samples:
            key = tuple(genotype for genotype in sample.genotypes if genotype is not None)
            groups.setdefault(key, []).append(sample)
        chosen: list[SampleGenotype] = []
        for group in groups.values():
            chosen.append(min(group, key=lambda sample: (_norm_sample_id(sample.sample_id), sample.source_row, sample.sample_id)))
        chosen.sort(key=lambda sample: sample.source_row)
        duplicate_counts[population_id] = len(population_samples) - len(chosen)
        if len(chosen) < 2:
            raise NonEstimable(
                "non_estimable_insufficient_clone_corrected_genets",
                f"{population_id} retains {len(chosen)} complete unique MLGs",
            )
        retained[population_id] = tuple(chosen)
    qc = {
        "n_raw_samples": len(samples),
        "n_complete_samples": len(complete),
        "n_incomplete_samples": len(samples) - len(complete),
        "complete_ramets_by_population": raw_complete_counts,
        "clone_duplicates_removed_by_population": duplicate_counts,
        "unique_complete_mlg_by_population": {population_id: len(retained[population_id]) for population_id in population_order},
        "n_unique_complete_mlg_total": sum(len(retained[population_id]) for population_id in population_order),
    }
    return retained, qc


def _population_loci(
    by_population: Mapping[str, Sequence[SampleGenotype]],
    *,
    population_order: Sequence[str],
    locus_order: Sequence[str],
) -> dict[str, dict[str, list[tuple[int, int]]]]:
    output: dict[str, dict[str, list[tuple[int, int]]]] = {}
    for population_id in population_order:
        records = by_population[population_id]
        output[population_id] = {}
        for locus_index, locus in enumerate(locus_order):
            values: list[tuple[int, int]] = []
            for record in records:
                genotype = record.genotypes[locus_index]
                if genotype is None:
                    raise RuntimeError("complete-case accounting drift")
                values.append(genotype)
            output[population_id][locus] = values
    return output


def validate_primary_theta(theta: float, pair_id: str) -> float:
    value = float(theta)
    if not math.isfinite(value) or value < 0.0 or value >= 1.0:
        raise NonEstimable("non_estimable_linearized_fst_domain", f"{pair_id} theta={value!r}")
    return value


def pairwise_theta_rows(
    by_population: Mapping[str, Sequence[SampleGenotype]],
    *,
    population_order: Sequence[str] = POPULATION_ORDER,
    locus_order: Sequence[str] = LOCUS_ORDER,
    enforce_primary_domain: bool,
) -> tuple[list[dict[str, object]], np.ndarray]:
    populations = _population_loci(
        by_population,
        population_order=population_order,
        locus_order=locus_order,
    )
    results = all_pairwise_theta(
        populations,
        population_order=population_order,
        locus_order=locus_order,
    )
    n = len(population_order)
    matrix = np.zeros((n, n), dtype=float)
    index = {population_id: i for i, population_id in enumerate(population_order)}
    rows: list[dict[str, object]] = []
    for result in results:
        theta = float(result.theta)
        if enforce_primary_domain:
            theta = validate_primary_theta(theta, f"{result.population_a}-{result.population_b}")
        linearized = theta / (1.0 - theta) if math.isfinite(theta) and theta < 1.0 else None
        if enforce_primary_domain and (linearized is None or not math.isfinite(linearized)):
            raise NonEstimable("non_estimable_linearized_fst_domain", f"non-finite transform {result.population_a}-{result.population_b}")
        i = index[result.population_a]
        j = index[result.population_b]
        matrix[i, j] = matrix[j, i] = theta
        rows.append(
            {
                "population_a": result.population_a,
                "population_b": result.population_b,
                "theta": theta,
                "linearized_fst": linearized,
                "n_usable_loci": result.n_usable_loci,
                "n_complete_genotypes_a": result.n_complete_genotypes_a,
                "n_complete_genotypes_b": result.n_complete_genotypes_b,
            }
        )
    expected_pairs = n * (n - 1) // 2
    if len(rows) != expected_pairs:
        raise RuntimeError("pair count drift")
    return rows, matrix


def ramet_population_map(
    samples: Sequence[SampleGenotype],
    *,
    population_order: Sequence[str] = POPULATION_ORDER,
) -> dict[str, tuple[SampleGenotype, ...]]:
    complete = complete_samples(samples)
    result = {
        population_id: tuple(sample for sample in complete if sample.population_id == population_id)
        for population_id in population_order
    }
    if any(len(result[population_id]) < 2 for population_id in population_order):
        raise NonEstimable("non_estimable_ramet_sensitivity", "a population has fewer than two complete ramets")
    return result


def load_predictor_matrices(
    path: Path = STAGE2_PREDICTORS,
    *,
    population_order: Sequence[str] = POPULATION_ORDER,
) -> tuple[dict[str, np.ndarray], np.ndarray, np.ndarray]:
    if path == STAGE2_PREDICTORS and _sha256(path) != EXPECTED_STAGE2_PREDICTORS_SHA256:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "predictor CSV SHA drifted")
    n = len(population_order)
    index = {population_id: i for i, population_id in enumerate(population_order)}
    reference = {candidate: np.zeros((n, n), dtype=float) for candidate in CANDIDATE_IDS}
    eog = np.zeros((n, n), dtype=float)
    disconnected = np.zeros((n, n), dtype=bool)
    seen: set[tuple[str, str]] = set()
    with path.open("r", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            a = str(row["population_a"])
            b = str(row["population_b"])
            if a not in index or b not in index or a == b:
                raise NonEstimable("non_estimable_stage2_predictor_drift", f"bad predictor pair {a}-{b}")
            key = tuple(sorted((a, b)))
            if key in seen:
                raise NonEstimable("non_estimable_stage2_predictor_drift", f"duplicate predictor pair {a}-{b}")
            seen.add(key)
            i, j = index[a], index[b]
            for candidate in CANDIDATE_IDS:
                value = float(row[candidate])
                reference[candidate][i, j] = reference[candidate][j, i] = value
            value = float(row["eog_continuous_distance"])
            eog[i, j] = eog[j, i] = value
            flag = bool(int(row["eog_disconnected"]))
            disconnected[i, j] = disconnected[j, i] = flag
    if len(seen) != n * (n - 1) // 2:
        raise NonEstimable("non_estimable_stage2_predictor_drift", "incomplete predictor pair set")
    return reference, eog, disconnected


def fixed_geographic_pooled_mse(
    response_theta: np.ndarray,
    geographic: np.ndarray,
    *,
    population_order: Sequence[str] = POPULATION_ORDER,
    ridge_penalty: float = RIDGE_PENALTY,
) -> float:
    n = len(population_order)
    pairs = np.asarray([(i, j) for i in range(n) for j in range(i + 1, n)], dtype=int)
    raw = np.asarray([response_theta[i, j] for i, j in pairs], dtype=float)
    if np.any((raw < 0.0) | (raw >= 1.0) | ~np.isfinite(raw)):
        raise NonEstimable("non_estimable_linearized_fst_domain", "fixed geographic response domain failure")
    y = raw / (1.0 - raw)
    x = np.asarray([geographic[i, j] for i, j in pairs], dtype=float)[:, None]
    squared: list[float] = []
    for held_out in range(n):
        test = (pairs[:, 0] == held_out) | (pairs[:, 1] == held_out)
        train = ~test
        model = _fit_ridge(x[train], y[train], ridge_penalty)
        prediction = _predict_ridge(model, x[test])
        squared.extend(np.square(y[test] - prediction).tolist())
    return float(np.mean(squared))


def population_bootstrap_interval(
    delta_mse: Sequence[float],
    *,
    seed: int = BOOTSTRAP_SEED,
    resamples: int = BOOTSTRAP_RESAMPLES,
) -> tuple[float, float]:
    values = np.asarray(delta_mse, dtype=float)
    if len(values) == 0 or not np.isfinite(values).all():
        raise ValueError("delta_mse must be finite and non-empty")
    rng = np.random.default_rng(seed)
    draws = rng.integers(0, len(values), size=(resamples, len(values)))
    means = np.mean(values[draws], axis=1)
    low, high = np.percentile(means, [2.5, 97.5], method="linear")
    return float(low), float(high)


def _nested_payload(result) -> dict[str, object]:
    return {
        "node_ids": list(result.node_ids),
        "candidate_ids": list(result.candidate_ids),
        "config": asdict(result.config),
        "n_observed_pairs": result.n_observed_pairs,
        "folds": [
            {
                **{k: v for k, v in asdict(fold).items() if k != "inner_scores"},
                "inner_scores": [asdict(score) for score in fold.inner_scores],
            }
            for fold in result.folds
        ],
        "selection_counts": [list(item) for item in result.selection_counts],
        "baseline_pooled_mse": result.baseline_pooled_mse,
        "augmented_pooled_mse": result.augmented_pooled_mse,
        "pooled_delta_mse": result.pooled_delta_mse,
        "baseline_pooled_mae": result.baseline_pooled_mae,
        "augmented_pooled_mae": result.augmented_pooled_mae,
        "pooled_delta_mae": result.pooled_delta_mae,
        "mean_population_delta_mse": result.mean_population_delta_mse,
        "mean_population_delta_mae": result.mean_population_delta_mae,
        "fingerprint": result.fingerprint,
    }


def evaluate(raw_path: Path) -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    frozen = validate_archived_inputs()
    samples = parse_genotypes(raw_path)
    clone_corrected, qc = clone_correct(samples)
    primary_rows, response = pairwise_theta_rows(clone_corrected, enforce_primary_domain=True)

    references, eog, disconnected = load_predictor_matrices()
    config = GeneticValidationConfig(response_transform="linearized_fst", ridge_penalty=RIDGE_PENALTY)
    nested = evaluate_nested_genetic_reference_selection(
        POPULATION_ORDER,
        response,
        references,
        eog,
        disconnected,
        config=config,
    )
    geographic_mse = fixed_geographic_pooled_mse(response, references["geographic"])
    deltas = [fold.delta_mse for fold in nested.folds]
    bootstrap_low, bootstrap_high = population_bootstrap_interval(deltas)

    selector_competitive = nested.baseline_pooled_mse <= geographic_mse
    mean_improves = nested.mean_population_delta_mse < 0.0
    interval_below_zero = bootstrap_high < 0.0
    if not selector_competitive:
        status = "indeterminate_selector_reference_failure"
        promotion_go = False
    elif mean_improves and interval_below_zero:
        status = "promotion_go"
        promotion_go = True
    else:
        status = "null_or_adverse_eog_increment"
        promotion_go = False

    ramets = ramet_population_map(samples)
    sensitivity_rows, _ = pairwise_theta_rows(ramets, enforce_primary_domain=False)

    result: dict[str, object] = {
        "status": status,
        "promotion_go": promotion_go,
        "study_doi": "10.1111/mec.13966",
        "dryad_doi": "10.5061/dryad.404rm",
        "raw_genetic_sha256": EXPECTED_RAW_SHA256,
        "stage2_fingerprint": EXPECTED_STAGE2_FINGERPRINT,
        "stage3a_schema_fingerprint": EXPECTED_STAGE3A_SCHEMA_FINGERPRINT,
        "population_order": list(POPULATION_ORDER),
        "locus_order": list(LOCUS_ORDER),
        "n_primary_pairs": len(primary_rows),
        "qc": qc,
        "nested": _nested_payload(nested),
        "fixed_geographic_pooled_mse": geographic_mse,
        "selector_competitive_vs_geographic": selector_competitive,
        "mean_outer_population_delta_mse": nested.mean_population_delta_mse,
        "population_bootstrap": {
            "seed": BOOTSTRAP_SEED,
            "resamples": BOOTSTRAP_RESAMPLES,
            "percentile_method": "numpy.percentile method=linear",
            "lower_95": bootstrap_low,
            "upper_95": bootstrap_high,
        },
        "promotion_conditions": {
            "all_17_populations_and_136_pairs_estimable": True,
            "nested_selected_baseline_pooled_mse_lte_geographic": selector_competitive,
            "mean_outer_population_delta_mse_lt_zero": mean_improves,
            "bootstrap_upper_95_lt_zero": interval_below_zero,
        },
        "ramet_sensitivity_promotional": False,
        "frozen_stage2_manifest": {
            "candidate_family_fingerprint": frozen["stage2"]["candidate_family_fingerprint"],
            "graph_fingerprint": frozen["stage2"]["graph_fingerprint"],
            "operator_fingerprint": frozen["stage2"]["operator_fingerprint"],
            "connectivity_fingerprint": frozen["stage2"]["connectivity_fingerprint"],
        },
        "claim_boundary": (
            "Symmetric clone-corrected microsatellite FST tests added held-out information from frozen symmetric "
            "EOG exact-eventual reachability beyond a prospectively nested-selected conventional symmetric reference. "
            "It does not validate directional migration, asymmetric dispersal, colonization probability or calendar time."
        ),
    }
    result["fingerprint"] = _canonical_sha256(
        {"result": result, "primary_pairs": primary_rows, "ramet_sensitivity_pairs": sensitivity_rows}
    )
    return result, primary_rows, sensitivity_rows


def _write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError("cannot write empty CSV")
    fieldnames = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    try:
        result, primary_rows, sensitivity_rows = evaluate(args.input)
    except NonEstimable as exc:
        result = {
            "status": exc.status,
            "promotion_go": False,
            "detail": exc.detail,
            "raw_genetic_sha256": EXPECTED_RAW_SHA256,
            "stage2_fingerprint": EXPECTED_STAGE2_FINGERPRINT,
            "stage3a_schema_fingerprint": EXPECTED_STAGE3A_SCHEMA_FINGERPRINT,
        }
        result["fingerprint"] = _canonical_sha256(result)
        (args.output_dir / "result.json").write_text(
            json.dumps(result, indent=2, sort_keys=True, allow_nan=False) + "\n",
            encoding="utf-8",
        )
        print(json.dumps(result, indent=2, sort_keys=True, allow_nan=False))
        return

    _write_csv(args.output_dir / "primary_pairwise_fst.csv", primary_rows)
    _write_csv(args.output_dir / "ramet_sensitivity_pairwise_fst.csv", sensitivity_rows)
    (args.output_dir / "result.json").write_text(
        json.dumps(result, indent=2, sort_keys=True, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(result, indent=2, sort_keys=True, allow_nan=False))


if __name__ == "__main__":
    main()
