"""Integrity-gated schema audit for the Czech dry-grassland benchmark.

This module deliberately stops before any EOG graph, support model, species-level
performance statistic, or outcome-based filtering is computed.  Workbook contents
are inspected only after the bytes match the source SHA-256 frozen in PR #95.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import re
from typing import Any

EXPECTED_NAME = "Belinchón.et.al.Oikos2020.data.xlsx"
EXPECTED_SIZE = 75727
EXPECTED_SHA256 = "ddc63118750c4685ec4ab35560299de312c57a8a120a71b27f31d94193e37867"

COORD_TOKENS = {"lat", "latitude", "lon", "long", "longitude", "x", "y", "coord", "utm", "easting", "northing"}
PATCH_TOKENS = {"patch", "site", "plot", "locality", "id"}
ABIOTIC_TOKENS = {"ph", "soil", "slope", "aspect", "temperature", "temp", "moisture", "nutrient", "nitrogen", "phosphorus", "light", "productivity"}
LANDSCAPE_TOKENS = {"area", "size", "isolation", "distance", "connect", "landscape", "grassland", "historical", "history", "past", "present"}
DARK_TOKENS = {"dark", "absent", "pool", "potential"}
OCC_TOKENS = {"occur", "occurrence", "presence", "observed", "community", "species", "spp"}


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def verify(path: Path) -> dict[str, Any]:
    size = path.stat().st_size
    digest = _sha256(path)
    ok = size == EXPECTED_SIZE and digest == EXPECTED_SHA256
    result = {
        "path": str(path),
        "name": path.name,
        "size": size,
        "sha256": digest,
        "expected_size": EXPECTED_SIZE,
        "expected_sha256": EXPECTED_SHA256,
        "integrity_ok": ok,
    }
    if not ok:
        raise ValueError(
            "Czech workbook failed frozen-source integrity gate: "
            f"size={size} sha256={digest}"
        )
    return result


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _hits(headers: list[str], tokens: set[str]) -> list[str]:
    found: list[str] = []
    for header in headers:
        words = set(_norm(header).split())
        if words & tokens or any(token in _norm(header) for token in tokens if len(token) >= 4):
            found.append(header)
    return found


def _sheet_audit(ws) -> dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    headers: list[str] = []
    if max_row:
        headers = ["" if ws.cell(1, c).value is None else str(ws.cell(1, c).value) for c in range(1, max_col + 1)]

    # Structural summaries only.  No biological response is aggregated or scored.
    numeric_columns: list[str] = []
    binary_columns: list[str] = []
    nonempty_counts: dict[str, int] = {}
    unique_counts: dict[str, int] = {}
    sample_limit = min(max_row, 273)
    for c, header in enumerate(headers, start=1):
        label = header or f"column_{c}"
        values = [ws.cell(r, c).value for r in range(2, sample_limit + 1)]
        values = [v for v in values if v not in (None, "")]
        nonempty_counts[label] = len(values)
        unique_counts[label] = len({str(v) for v in values})
        if values:
            try:
                nums = [float(v) for v in values]
                numeric_columns.append(label)
                if set(nums) <= {0.0, 1.0}:
                    binary_columns.append(label)
            except (TypeError, ValueError):
                pass

    return {
        "sheet": ws.title,
        "n_rows_including_header": max_row,
        "n_columns": max_col,
        "headers": headers,
        "candidate_patch_id_columns": _hits(headers, PATCH_TOKENS),
        "candidate_coordinate_columns": _hits(headers, COORD_TOKENS),
        "candidate_abiotic_columns": _hits(headers, ABIOTIC_TOKENS),
        "candidate_landscape_columns": _hits(headers, LANDSCAPE_TOKENS),
        "candidate_observed_occurrence_columns": _hits(headers, OCC_TOKENS),
        "candidate_dark_diversity_columns": _hits(headers, DARK_TOKENS),
        "numeric_columns": numeric_columns,
        "binary_columns": binary_columns,
        "nonempty_counts_first_272_rows": nonempty_counts,
        "unique_counts_first_272_rows": unique_counts,
    }


def audit(workbook: Path, output: Path) -> dict[str, Any]:
    integrity = verify(workbook)
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Czech workbook schema audit") from exc

    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    sheets = [_sheet_audit(book[name]) for name in book.sheetnames]
    report = {
        "status": "preoutcome_schema_audit",
        "dataset": "Belinchón, Hemrová & Münzbergová Czech dry-grassland benchmark",
        "doi": "10.5061/dryad.jh9w0vt8f",
        "integrity": integrity,
        "sheet_names": book.sheetnames,
        "sheets": sheets,
        "geometry_gate": {
            "coordinate_like_columns_detected": any(s["candidate_coordinate_columns"] for s in sheets),
            "rule": "Do not substitute published scalar isolation for graph geometry. A graph benchmark requires coordinates or explicit pairwise geometry in the verified workbook/source materials.",
        },
        "scientific_boundary": "schema/geometry audit only; no support model, taxon eligibility filtering, graph, EOG statistic, concordance, bottleneck score, or model selection computed",
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return report


def blocked_report(output: Path, reason: str) -> dict[str, Any]:
    report = {
        "status": "blocked_before_schema",
        "doi": "10.5061/dryad.jh9w0vt8f",
        "expected_file": EXPECTED_NAME,
        "expected_size": EXPECTED_SIZE,
        "expected_sha256": EXPECTED_SHA256,
        "reason": reason,
        "next_gate": "Obtain the workbook bytes, verify frozen size/SHA-256, then run this schema auditor unchanged.",
        "scientific_boundary": "no workbook contents or EOG outcomes inspected",
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--blocked-reason")
    args = parser.parse_args()
    if args.workbook:
        print(json.dumps(audit(args.workbook, args.output), indent=2, ensure_ascii=False))
    else:
        reason = args.blocked_reason or "verified workbook bytes were not supplied"
        print(json.dumps(blocked_report(args.output, reason), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
