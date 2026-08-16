#!/usr/bin/env python3
"""Freeze the 22 response-blind SIVFLORA island nodes.

Only the ``islands_data`` worksheet is read. The species-incidence worksheet is never
opened by this stage. The explicit primary island rows (non-null ``Id``) define the
geographic node universe fixed by pre-outcome amendment 001.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SOURCE_SHA256 = "6c9715e5a3b39942a9c9c9e364a85bb7fa9024697cb19c9d82dac30920935bdf"
EXPECTED_NODE_COUNT = 22


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse_dms(value: object, *, latitude: bool) -> float:
    text = str(value or "").strip().upper().replace("\u00a0", " ")
    text = text.replace("′", "'").replace("’", "'").replace("″", '"').replace("“", '"').replace("”", '"')
    hemisphere_match = re.search(r"([NSEW])\s*$", text)
    if hemisphere_match is None:
        raise ValueError(f"coordinate lacks hemisphere: {value!r}")
    hemisphere = hemisphere_match.group(1)
    if latitude and hemisphere not in {"N", "S"}:
        raise ValueError(f"latitude has invalid hemisphere: {value!r}")
    if not latitude and hemisphere not in {"E", "W"}:
        raise ValueError(f"longitude has invalid hemisphere: {value!r}")

    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    if not 1 <= len(numbers) <= 3:
        raise ValueError(f"unable to parse DMS coordinate: {value!r}")
    degrees = numbers[0]
    minutes = numbers[1] if len(numbers) >= 2 else 0.0
    seconds = numbers[2] if len(numbers) >= 3 else 0.0
    if minutes >= 60 or seconds >= 60:
        raise ValueError(f"invalid DMS minutes/seconds: {value!r}")
    coordinate = degrees + minutes / 60.0 + seconds / 3600.0
    if hemisphere in {"S", "W"}:
        coordinate *= -1.0
    limit = 90.0 if latitude else 180.0
    if not math.isfinite(coordinate) or abs(coordinate) > limit:
        raise ValueError(f"coordinate outside valid range: {value!r}")
    return coordinate


def freeze_nodes(source_xlsx: Path, output_csv: Path, output_manifest: Path) -> dict[str, object]:
    actual_source_sha = _sha256(source_xlsx)
    if actual_source_sha != EXPECTED_SOURCE_SHA256:
        raise ValueError(
            f"SIVFLORA source SHA-256 mismatch: expected {EXPECTED_SOURCE_SHA256}, got {actual_source_sha}"
        )

    workbook = load_workbook(source_xlsx, read_only=True, data_only=True)
    if "islands_data" not in workbook.sheetnames:
        raise ValueError("frozen workbook lacks islands_data worksheet")
    sheet = workbook["islands_data"]
    iterator = sheet.iter_rows(values_only=True)
    header = tuple(next(iterator))
    required = (
        "Id",
        "Acronym",
        "Island-Archipelago",
        "Latitude",
        "Longitude",
        "Surface (km2)",
    )
    index = {name: header.index(name) for name in required if name in header}
    if set(index) != set(required):
        missing = sorted(set(required) - set(index))
        raise ValueError(f"islands_data missing required columns: {missing}")

    nodes: list[dict[str, object]] = []
    for values in iterator:
        raw_id = values[index["Id"]]
        if raw_id is None:
            continue
        island_id = int(raw_id)
        acronym = str(values[index["Acronym"]] or "").strip()
        name = str(values[index["Island-Archipelago"]] or "").strip()
        latitude_raw = values[index["Latitude"]]
        longitude_raw = values[index["Longitude"]]
        if not acronym or not name:
            raise ValueError(f"primary island row {island_id} lacks acronym/name")
        nodes.append(
            {
                "island_id": island_id,
                "acronym": acronym,
                "node_name": name,
                "latitude": parse_dms(latitude_raw, latitude=True),
                "longitude": parse_dms(longitude_raw, latitude=False),
                "latitude_source": str(latitude_raw),
                "longitude_source": str(longitude_raw),
                "surface_km2_source": str(values[index["Surface (km2)"]]),
            }
        )
    workbook.close()

    nodes.sort(key=lambda row: int(row["island_id"]))
    ids = [int(row["island_id"]) for row in nodes]
    names = [str(row["node_name"]) for row in nodes]
    acronyms = [str(row["acronym"]) for row in nodes]
    if len(nodes) != EXPECTED_NODE_COUNT:
        raise ValueError(f"expected {EXPECTED_NODE_COUNT} primary island nodes, got {len(nodes)}")
    if ids != list(range(1, EXPECTED_NODE_COUNT + 1)):
        raise ValueError(f"primary island IDs must be exactly 1..{EXPECTED_NODE_COUNT}, got {ids}")
    if len(set(names)) != EXPECTED_NODE_COUNT or len(set(acronyms)) != EXPECTED_NODE_COUNT:
        raise ValueError("primary island names/acronyms must be unique")

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fields = (
        "island_id",
        "acronym",
        "node_name",
        "latitude",
        "longitude",
        "latitude_source",
        "longitude_source",
        "surface_km2_source",
    )
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in nodes:
            serial = dict(row)
            serial["latitude"] = f"{float(row['latitude']):.10f}"
            serial["longitude"] = f"{float(row['longitude']):.10f}"
            writer.writerow(serial)

    manifest = {
        "status": "preoutcome_node_freeze",
        "source_sha256": actual_source_sha,
        "source_sheet_read": "islands_data",
        "species_incidence_sheet_opened": False,
        "outcome_statistics_computed": False,
        "n_nodes": len(nodes),
        "node_rule": "non-null Id primary rows; explicit island-level DMS coordinate",
        "species_weighting_used": False,
        "fallback_or_locality_substitution_used": False,
        "nodes_csv_sha256": _sha256(output_csv),
    }
    output_manifest.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-xlsx", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-manifest", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(freeze_nodes(args.source_xlsx, args.output_csv, args.output_manifest), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
