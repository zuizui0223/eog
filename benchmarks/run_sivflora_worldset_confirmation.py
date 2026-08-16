#!/usr/bin/env python3
"""Run the once-only frozen SIVFLORA explicit-world confirmation.

Scientific design is frozen in ``sivflora_confirmation_contract.json`` plus pre-outcome
amendments 001/002. This runner must not tune worlds, climate variables, thresholds,
regularization, or comparator definitions from the observed outcome.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import warnings
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from sklearn.exceptions import ConvergenceWarning
from sklearn.linear_model import LogisticRegression


EXPECTED_SOURCE_SHA256 = "6c9715e5a3b39942a9c9c9e364a85bb7fa9024697cb19c9d82dac30920935bdf"
VARIABLES = ("bio1", "bio5", "bio6", "bio12", "bio15")
EARTH_RADIUS_KM = 6371.0088
MIN_TRAIN_POS = 4
MIN_TRAIN_NEG = 4
MIN_PAIRS_PER_ISLAND = 20
MIN_EVALUABLE_ISLANDS = 15
BOOTSTRAP_REPLICATES = 10000
BOOTSTRAP_SEED = 20260816
CLIP = 1e-15


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * EARTH_RADIUS_KM * math.asin(min(1.0, math.sqrt(a)))


def _pairwise_geo(rows: list[dict[str, str]]) -> np.ndarray:
    n = len(rows)
    out = np.zeros((n, n), dtype=float)
    for i in range(n):
        for j in range(i + 1, n):
            value = _haversine_km(
                float(rows[i]["latitude"]), float(rows[i]["longitude"]),
                float(rows[j]["latitude"]), float(rows[j]["longitude"]),
            )
            out[i, j] = out[j, i] = value
    return out


def _pairwise_env(rows: list[dict[str, str]], prefix: str) -> np.ndarray:
    data = np.asarray([[float(row[f"{prefix}_{v}"]) for v in VARIABLES] for row in rows], dtype=float)
    if not np.isfinite(data).all():
        raise ValueError(f"{prefix} climate contains non-finite values")
    mean = data.mean(axis=0)
    sd = data.std(axis=0, ddof=1)
    if np.any(~np.isfinite(sd)) or np.any(sd <= 0):
        raise ValueError(f"{prefix} climate has zero/non-finite sample SD")
    z = (data - mean) / sd
    diff = z[:, None, :] - z[None, :, :]
    return np.sqrt(np.sum(diff * diff, axis=2))


def _read_climate(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != 22:
        raise ValueError(f"frozen climate table must contain 22 nodes, got {len(rows)}")
    if tuple(row["island_id"] for row in rows) != tuple(str(i) for i in range(1, 23)):
        raise ValueError("frozen climate node order must be island IDs 1..22")
    return rows


def _load_world_components(worlds_path: Path) -> tuple[list[str], list[str], np.ndarray]:
    payload = json.loads(worlds_path.read_text(encoding="utf-8"))
    if payload.get("world_count") != 20 or len(payload.get("worlds", [])) != 20:
        raise ValueError("confirmation requires exactly the frozen 20-world universe")
    node_order = [str(value) for value in payload["node_order"]]
    if node_order != [str(i) for i in range(1, 23)]:
        raise ValueError("world universe node order differs from frozen 1..22 order")
    world_ids: list[str] = []
    families: list[str] = []
    comp = np.full((20, 22), -1, dtype=int)
    for w, world in enumerate(payload["worlds"]):
        world_ids.append(str(world["world_id"]))
        families.append(str(world["family"]))
        for component_id, nodes in enumerate(world["components"]):
            for node in nodes:
                index = node_order.index(str(node))
                if comp[w, index] != -1:
                    raise ValueError("node occurs in multiple components within one world")
                comp[w, index] = component_id
    if np.any(comp < 0) or len(set(world_ids)) != 20:
        raise ValueError("invalid frozen world component coverage")
    expected_families = {
        "geography_only": 4,
        "chelsa_q50": 4,
        "chelsa_q75": 4,
        "worldclim_q50": 4,
        "worldclim_q75": 4,
    }
    actual = {name: families.count(name) for name in expected_families}
    if actual != expected_families:
        raise ValueError(f"world family counts changed: {actual}")
    return world_ids, families, comp


def pair_state(has_natural: bool, has_other_record: bool, has_any_record: bool) -> int | None:
    """Return 1 natural positive, 0 catalogue non-record, None excluded/ambiguous."""
    if has_natural and not has_other_record:
        return 1
    if not has_any_record:
        return 0
    return None


def _load_incidence(source_xlsx: Path, schema_audit: Path, node_rows: list[dict[str, str]]) -> tuple[list[str], dict[tuple[str, int], int | None], dict[str, object]]:
    if _sha256(source_xlsx) != EXPECTED_SOURCE_SHA256:
        raise ValueError("SIVFLORA source SHA-256 differs from the frozen source")
    audit = json.loads(schema_audit.read_text(encoding="utf-8"))
    name_map = {str(k): str(v) for k, v in audit["main_to_island_metadata_name"].items()}
    node_by_name = {row["node_name"]: int(row["island_id"]) - 1 for row in node_rows}
    if set(name_map.values()) != set(node_by_name):
        raise ValueError("schema name map no longer covers exactly the 22 frozen nodes")

    workbook = load_workbook(source_xlsx, read_only=True, data_only=True)
    sheet = workbook["sivflora"]
    iterator = sheet.iter_rows(values_only=True)
    header = tuple(next(iterator))
    required = ("island", "reviewed_name", "establishment_means", "gbif_rank")
    index = {name: header.index(name) for name in required if name in header}
    if set(index) != set(required):
        raise ValueError(f"sivflora sheet missing required fields: {sorted(set(required)-set(index))}")

    records: dict[tuple[str, int], dict[str, bool]] = defaultdict(
        lambda: {"natural": False, "other": False, "any": False}
    )
    species_seen: set[str] = set()
    species_rows = 0
    for values in iterator:
        rank = str(values[index["gbif_rank"]] or "").strip().upper()
        if rank != "SPECIES":
            continue
        taxon = str(values[index["reviewed_name"]] or "").strip()
        island_main = str(values[index["island"]] or "").strip()
        if not taxon or not island_main:
            raise ValueError("species-rank row lacks reviewed_name or island")
        if island_main not in name_map:
            raise ValueError(f"unmapped SIVFLORA island name: {island_main!r}")
        node_name = name_map[island_main]
        node = node_by_name[node_name]
        establishment = str(values[index["establishment_means"]] or "").strip().casefold()
        natural = establishment in {"native", "endemic"}
        key = (taxon, node)
        records[key]["any"] = True
        if natural:
            records[key]["natural"] = True
        else:
            records[key]["other"] = True
        species_seen.add(taxon)
        species_rows += 1
    workbook.close()

    states: dict[tuple[str, int], int | None] = {}
    natural_taxa: list[str] = []
    conflict_pairs = 0
    excluded_record_pairs = 0
    for taxon in sorted(species_seen):
        row_states: list[int | None] = []
        for node in range(22):
            flags = records.get((taxon, node))
            if flags is None:
                state = 0
            else:
                state = pair_state(flags["natural"], flags["other"], flags["any"])
                if flags["natural"] and flags["other"]:
                    conflict_pairs += 1
                elif state is None:
                    excluded_record_pairs += 1
            states[(taxon, node)] = state
            row_states.append(state)
        if any(value == 1 for value in row_states):
            natural_taxa.append(taxon)

    # Remove taxa that are never natural anywhere from the evaluation universe.
    states = {key: value for key, value in states.items() if key[0] in set(natural_taxa)}
    summary = {
        "species_rank_source_rows": species_rows,
        "species_rank_reviewed_names": len(species_seen),
        "taxa_with_at_least_one_natural_or_endemic_island": len(natural_taxa),
        "conflicting_natural_other_taxon_island_pairs": conflict_pairs,
        "other_record_taxon_island_pairs_excluded": excluded_record_pairs,
    }
    return natural_taxa, states, summary


def _world_bits(target: int, anchors: list[int], components: np.ndarray) -> np.ndarray:
    if not anchors:
        raise ValueError("reachability feature requires at least one positive anchor")
    anchor_array = np.asarray(anchors, dtype=int)
    return np.asarray(
        [np.any(components[w, anchor_array] == components[w, target]) for w in range(components.shape[0])],
        dtype=float,
    )


def _family_counts(bits: np.ndarray, families: list[str]) -> tuple[float, float, float, float, float]:
    names = ("geography_only", "chelsa_q50", "chelsa_q75", "worldclim_q50", "worldclim_q75")
    return tuple(float(sum(bits[i] for i, family in enumerate(families) if family == name)) for name in names)


def _feature(
    target: int,
    reference_nodes: list[int],
    taxon: str,
    states: dict[tuple[str, int], int | None],
    geo: np.ndarray,
    chelsa: np.ndarray,
    worldclim: np.ndarray,
    components: np.ndarray,
    families: list[str],
) -> tuple[np.ndarray, np.ndarray]:
    evaluable = [node for node in reference_nodes if states[(taxon, node)] is not None]
    anchors = [node for node in reference_nodes if states[(taxon, node)] == 1]
    if not evaluable or not anchors:
        raise ValueError("feature reference set lacks evaluable rows or positive anchors")
    prevalence = len(anchors) / len(evaluable)
    nearest_geo = min(float(geo[target, node]) for node in anchors)
    nearest_chelsa = min(float(chelsa[target, node]) for node in anchors)
    nearest_worldclim = min(float(worldclim[target, node]) for node in anchors)
    bits = _world_bits(target, anchors, components)
    counts = _family_counts(bits, families)
    compressed = np.asarray(
        [
            prevalence,
            math.log1p(nearest_geo),
            nearest_chelsa,
            nearest_worldclim,
            float(bits.mean()),
            *counts,
        ],
        dtype=float,
    )
    if compressed.shape != (10,) or bits.shape != (20,) or not np.isfinite(compressed).all():
        raise AssertionError("frozen feature shape/value contract failed")
    return compressed, bits


def _scale(train: np.ndarray, test: np.ndarray, n_scale: int) -> tuple[np.ndarray, np.ndarray]:
    train_out = np.asarray(train, dtype=float).copy()
    test_out = np.asarray(test, dtype=float).copy()
    mean = train_out[:, :n_scale].mean(axis=0)
    sd = train_out[:, :n_scale].std(axis=0, ddof=0)
    safe = np.where(sd > 0, sd, 1.0)
    train_out[:, :n_scale] = (train_out[:, :n_scale] - mean) / safe
    test_out[:, :n_scale] = (test_out[:, :n_scale] - mean) / safe
    if np.any(sd == 0):
        train_out[:, :n_scale][:, sd == 0] = 0.0
        test_out[:, :n_scale][:, sd == 0] = 0.0
    return train_out, test_out


def _fit_predict(train: np.ndarray, y: np.ndarray, test: np.ndarray, n_scale: int) -> np.ndarray:
    if len(np.unique(y)) < 2:
        raise ValueError("outer training labels contain one class")
    x_train, x_test = _scale(train, test, n_scale)
    model = LogisticRegression(
        penalty="l2",
        C=1.0,
        solver="lbfgs",
        fit_intercept=True,
        class_weight=None,
        max_iter=5000,
        tol=1e-4,
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error", ConvergenceWarning)
        model.fit(x_train, y)
    return np.clip(model.predict_proba(x_test)[:, 1], CLIP, 1.0 - CLIP)


def _logloss(y: np.ndarray, p: np.ndarray) -> float:
    p = np.clip(np.asarray(p, dtype=float), CLIP, 1.0 - CLIP)
    y = np.asarray(y, dtype=float)
    return float(np.mean(-(y * np.log(p) + (1.0 - y) * np.log(1.0 - p))))


def _brier(y: np.ndarray, p: np.ndarray) -> float:
    return float(np.mean((np.asarray(p, dtype=float) - np.asarray(y, dtype=float)) ** 2))


def _identity_not_determined(rows: list[dict[str, object]]) -> tuple[bool, int, int]:
    signatures: dict[tuple[int, ...], set[str]] = defaultdict(set)
    for row in rows:
        compressed = tuple(int(row[name]) for name in (
            "geography_only_support_count",
            "chelsa_q50_support_count",
            "chelsa_q75_support_count",
            "worldclim_q50_support_count",
            "worldclim_q75_support_count",
        ))
        signatures[compressed].add(str(row["world_bits"]))
    collision_groups = sum(len(values) > 1 for values in signatures.values())
    collision_rows = sum(
        1
        for row in rows
        if len(signatures[tuple(int(row[name]) for name in (
            "geography_only_support_count",
            "chelsa_q50_support_count",
            "chelsa_q75_support_count",
            "worldclim_q50_support_count",
            "worldclim_q75_support_count",
        ))]) > 1
    )
    return collision_groups > 0, collision_groups, collision_rows


def run(
    source_xlsx: Path,
    nodes_csv: Path,
    climate_csv: Path,
    worlds_json: Path,
    schema_audit: Path,
    pair_output: Path,
    island_output: Path,
    result_output: Path,
) -> dict[str, object]:
    node_rows = _read_climate(climate_csv)
    nodes_source_rows = list(csv.DictReader(nodes_csv.open(newline="", encoding="utf-8")))
    if [(r["island_id"], r["node_name"]) for r in node_rows] != [(r["island_id"], r["node_name"]) for r in nodes_source_rows]:
        raise ValueError("nodes and climate tables do not share the same frozen node identity/order")

    world_ids, families, components = _load_world_components(worlds_json)
    geo = _pairwise_geo(node_rows)
    chelsa = _pairwise_env(node_rows, "chelsa")
    worldclim = _pairwise_env(node_rows, "worldclim")
    taxa, states, incidence_summary = _load_incidence(source_xlsx, schema_audit, node_rows)

    pair_rows: list[dict[str, object]] = []
    island_rows: list[dict[str, object]] = []
    fold_failures: list[dict[str, object]] = []

    for heldout in range(22):
        training_nodes = [node for node in range(22) if node != heldout]
        eligible_taxa: list[str] = []
        for taxon in taxa:
            positives = sum(states[(taxon, node)] == 1 for node in training_nodes)
            negatives = sum(states[(taxon, node)] == 0 for node in training_nodes)
            if positives >= MIN_TRAIN_POS and negatives >= MIN_TRAIN_NEG:
                eligible_taxa.append(taxon)

        train_comp: list[np.ndarray] = []
        train_bits: list[np.ndarray] = []
        train_y: list[int] = []
        for taxon in eligible_taxa:
            for target in training_nodes:
                label = states[(taxon, target)]
                if label is None:
                    continue
                reference = [node for node in training_nodes if node != target]
                compressed, bits = _feature(
                    target, reference, taxon, states, geo, chelsa, worldclim, components, families
                )
                train_comp.append(compressed)
                train_bits.append(bits)
                train_y.append(int(label))

        test_taxa = [taxon for taxon in eligible_taxa if states[(taxon, heldout)] is not None]
        if not train_y or not test_taxa:
            fold_failures.append({"heldout_island_id": heldout + 1, "reason": "no_training_or_test_rows"})
            continue

        test_comp: list[np.ndarray] = []
        test_bits: list[np.ndarray] = []
        test_y: list[int] = []
        for taxon in test_taxa:
            compressed, bits = _feature(
                heldout, training_nodes, taxon, states, geo, chelsa, worldclim, components, families
            )
            test_comp.append(compressed)
            test_bits.append(bits)
            test_y.append(int(states[(taxon, heldout)]))

        tc = np.vstack(train_comp)
        tb = np.vstack(train_bits)
        yc = np.asarray(train_y, dtype=int)
        xc = np.vstack(test_comp)
        xb = np.vstack(test_bits)
        y_test = np.asarray(test_y, dtype=int)

        try:
            p_r0 = _fit_predict(tc[:, :4], yc, xc[:, :4], 4)
            p_r1 = _fit_predict(tc[:, :5], yc, xc[:, :5], 5)
            p_r2 = _fit_predict(tc[:, :10], yc, xc[:, :10], 10)
            p_c = _fit_predict(np.hstack([tc[:, :10], tb]), yc, np.hstack([xc[:, :10], xb]), 10)
        except Exception as exc:
            fold_failures.append({"heldout_island_id": heldout + 1, "reason": f"model_fit_error:{type(exc).__name__}:{exc}"})
            continue

        for i, taxon in enumerate(test_taxa):
            counts = [int(v) for v in xc[i, 5:10]]
            pair_rows.append(
                {
                    "heldout_island_id": heldout + 1,
                    "heldout_island": node_rows[heldout]["node_name"],
                    "taxon": taxon,
                    "catalogued_natural_incidence": int(y_test[i]),
                    "p_R0": float(p_r0[i]),
                    "p_R1": float(p_r1[i]),
                    "p_R2": float(p_r2[i]),
                    "p_C_identity": float(p_c[i]),
                    "training_prevalence": float(xc[i, 0]),
                    "log1p_nearest_geo_km": float(xc[i, 1]),
                    "nearest_chelsa_env": float(xc[i, 2]),
                    "nearest_worldclim_env": float(xc[i, 3]),
                    "world_frequency": float(xc[i, 4]),
                    "geography_only_support_count": counts[0],
                    "chelsa_q50_support_count": counts[1],
                    "chelsa_q75_support_count": counts[2],
                    "worldclim_q50_support_count": counts[3],
                    "worldclim_q75_support_count": counts[4],
                    "world_bits": "".join(str(int(v)) for v in xb[i]),
                }
            )

        if len(test_taxa) >= MIN_PAIRS_PER_ISLAND:
            island_rows.append(
                {
                    "heldout_island_id": heldout + 1,
                    "heldout_island": node_rows[heldout]["node_name"],
                    "n_pairs": len(test_taxa),
                    "logloss_R0": _logloss(y_test, p_r0),
                    "logloss_R1": _logloss(y_test, p_r1),
                    "logloss_R2": _logloss(y_test, p_r2),
                    "logloss_C_identity": _logloss(y_test, p_c),
                    "brier_R0": _brier(y_test, p_r0),
                    "brier_R1": _brier(y_test, p_r1),
                    "brier_R2": _brier(y_test, p_r2),
                    "brier_C_identity": _brier(y_test, p_c),
                    "delta_identity": _logloss(y_test, p_c) - _logloss(y_test, p_r2),
                }
            )

    pair_output.parent.mkdir(parents=True, exist_ok=True)
    pair_fields = list(pair_rows[0]) if pair_rows else ["heldout_island_id"]
    with pair_output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=pair_fields)
        writer.writeheader()
        writer.writerows(pair_rows)

    island_fields = list(island_rows[0]) if island_rows else ["heldout_island_id"]
    with island_output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=island_fields)
        writer.writeheader()
        writer.writerows(island_rows)

    identity_extra, collision_groups, collision_rows = _identity_not_determined(pair_rows) if pair_rows else (False, 0, 0)
    coverage_ok = len(island_rows) >= MIN_EVALUABLE_ISLANDS
    result: dict[str, object] = {
        "design": "SIVFLORA independent explicit-world identity confirmation",
        "source_sha256": _sha256(source_xlsx),
        "nodes_sha256": _sha256(nodes_csv),
        "climate_sha256": _sha256(climate_csv),
        "worlds_sha256": _sha256(worlds_json),
        "world_ids": world_ids,
        "incidence_summary": incidence_summary,
        "n_pair_predictions": len(pair_rows),
        "n_evaluable_outer_islands": len(island_rows),
        "minimum_required_evaluable_outer_islands": MIN_EVALUABLE_ISLANDS,
        "coverage_ok": coverage_ok,
        "fold_failures": fold_failures,
        "identity_not_determined_by_R2_decomposition": identity_extra,
        "R2_decomposition_collision_groups": collision_groups,
        "rows_in_R2_decomposition_collision_groups": collision_rows,
        "pair_output_sha256": _sha256(pair_output),
        "island_output_sha256": _sha256(island_output),
    }

    if not coverage_ok:
        result.update(
            {
                "status": "indeterminate_non_estimable",
                "favourable": False,
                "reason": "fewer than 15 outer islands had at least 20 evaluable held-out taxon pairs",
            }
        )
    else:
        deltas = np.asarray([float(row["delta_identity"]) for row in island_rows], dtype=float)
        rng = np.random.default_rng(BOOTSTRAP_SEED)
        indices = rng.integers(0, len(deltas), size=(BOOTSTRAP_REPLICATES, len(deltas)))
        boot = deltas[indices].mean(axis=1)
        ci_low, ci_high = np.quantile(boot, [0.025, 0.975], method="linear")
        mean_r0 = float(np.mean([float(row["logloss_R0"]) for row in island_rows]))
        mean_r1 = float(np.mean([float(row["logloss_R1"]) for row in island_rows]))
        mean_r2 = float(np.mean([float(row["logloss_R2"]) for row in island_rows]))
        mean_c = float(np.mean([float(row["logloss_C_identity"]) for row in island_rows]))
        better_islands = int(np.sum(deltas < 0))
        conditions = {
            "mean_delta_identity_lt_zero": float(deltas.mean()) < 0,
            "bootstrap_upper_lt_zero": float(ci_high) < 0,
            "candidate_better_outer_islands_at_least_12": better_islands >= 12,
            "candidate_mean_not_worse_than_R1": mean_c <= mean_r1,
            "world_identity_not_deterministic_from_R2_decomposition": identity_extra,
        }
        favourable = all(conditions.values())
        result.update(
            {
                "status": "confirmed_added_value" if favourable else "no_confirmed_added_value",
                "favourable": favourable,
                "mean_island_macro_logloss_R0": mean_r0,
                "mean_island_macro_logloss_R1": mean_r1,
                "mean_island_macro_logloss_R2": mean_r2,
                "mean_island_macro_logloss_C_identity": mean_c,
                "mean_delta_identity_C_minus_R2": float(deltas.mean()),
                "bootstrap_95_low": float(ci_low),
                "bootstrap_95_high": float(ci_high),
                "C_better_than_R2_outer_islands": better_islands,
                "favourable_conditions": conditions,
                "bootstrap_replicates": BOOTSTRAP_REPLICATES,
                "bootstrap_seed": BOOTSTRAP_SEED,
            }
        )

    result_output.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-xlsx", type=Path, required=True)
    parser.add_argument("--nodes", type=Path, required=True)
    parser.add_argument("--climate", type=Path, required=True)
    parser.add_argument("--worlds", type=Path, required=True)
    parser.add_argument("--schema-audit", type=Path, required=True)
    parser.add_argument("--pair-output", type=Path, required=True)
    parser.add_argument("--island-output", type=Path, required=True)
    parser.add_argument("--result-output", type=Path, required=True)
    args = parser.parse_args()
    result = run(
        args.source_xlsx,
        args.nodes,
        args.climate,
        args.worlds,
        args.schema_audit,
        args.pair_output,
        args.island_output,
        args.result_output,
    )
    print(json.dumps(result, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
